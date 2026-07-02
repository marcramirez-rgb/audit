import os
import csv
import json
import time 
import urllib3
import zipfile
from io import BytesIO
from pathlib import Path
from datetime import datetime
import requests
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
from PIL import Image as PILImage, ImageDraw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

# --- GUI & Environment Imports ---
import tkinter as tk
from tkinter import filedialog
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Start the global execution benchmark timer
script_start_time = time.time()

# Suppress self-signed certificate warnings 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- 1. SET UP INFRASTRUCTURE PATHING & CREDENTIALS ---
USER = os.getenv("AXIS_USER", "root") 
PASSWORD = os.getenv("AXIS_PASSWORD")

# Secure Password Prompt Fallback
if not PASSWORD:
    print("[!] Security Notice: AXIS_PASSWORD environment variable not found.")
    PASSWORD = input("Please enter the camera password to continue: ").strip()
    if not PASSWORD:
        print("[-] Error: Password cannot be blank. Exiting.")
        exit()

STRICT_TIMEOUT = (3.05, 5.0)
THUMB_H = 360  
ROW_H   = 275  
MAX_ROWS_PER_FILE = 100 

downloads_path = Path.home() / "Downloads"
output_dir = downloads_path / "Camera_Reports_Batch"
output_dir.mkdir(parents=True, exist_ok=True)

# Lens-specific port configurations (Using Kingfisher Brand Hex Codes)
CAMERA_CONFIGS = [
    {"port": "5010", "position": "CENTER", "color": "E5F5F5"},  # Kingfisher Light
    {"port": "5015", "position": "LEFT",   "color": "00A19A"},  # Kingfisher Normal
    {"port": "5020", "position": "RIGHT",  "color": "00726E"}   # Kingfisher Dark
]

# --- 2. MODULAR WORKER FUNCTIONS ---

def fetch_axis_snapshot(session, ip, port, auth_strategies):
    """Fetches the background stream snapshot using the active session."""
    img_url = f"http://{ip}:{port}/axis-cgi/jpg/image.cgi?resolution=1280x720"
    for auth in auth_strategies:
        try:
            with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, stream=True, verify=False) as response:
                if response.status_code == 200:
                    return PILImage.open(BytesIO(response.content))
        except requests.exceptions.RequestException:
            continue
    return None

def fetch_axis_scenarios(session, ip, port, auth_strategies):
    """Fetches the Axis Object Analytics JSON payload."""
    control_url = f"http://{ip}:{port}/local/objectanalytics/control.cgi"
    payload = {"apiVersion": "1.2", "method": "getConfiguration"}
    
    for auth in auth_strategies:
        try:
            response = session.post(control_url, auth=auth, json=payload, timeout=STRICT_TIMEOUT, verify=False)
            if response.status_code == 200:
                return response.json(), control_url, None
            else:
                last_err = f"HTTP Error Status {response.status_code}"
        except Exception as e:
            last_err = str(e)
            continue
            
    return None, control_url, last_err

def parse_axis_analytics(json_data, img_w, img_h):
    """Parses Axis scenarios into structured rows and exact pixel coordinates."""
    scenarios = json_data.get("data", {}).get("scenarios", [])
    if not scenarios:
        return [{
            "is_placeholder": True, 
            "name": "No Scenarios Configured", 
            "type": "N/A", 
            "target": "No Analytics Configured", 
            "duration": "N/A", 
            "vertices": []
        }]
        
    parsed_rules = []
    for scenario in scenarios:
        if scenario.get("is_placeholder"):
            parsed_rules.append({
                "is_placeholder": True, 
                "name": "No Scenarios Configured", 
                "type": "N/A", 
                "target": "No Analytics Configured", 
                "duration": "N/A", 
                "vertices": []
            })
            continue

        rule_name = scenario.get("name", f"Scenario {scenario.get('id')}")
        loiter_time = ""
        triggers = scenario.get("triggers", [])
        
        for trigger_obj in triggers:
            for condition in trigger_obj.get("conditions", []):
                if condition.get("type") == "individualTimeInArea":
                    for d in condition.get("data", []):
                        if d.get("time"):
                            loiter_time = str(d["time"])
                            break
                            
        if not loiter_time:
            for f_obj in scenario.get("filters", []):
                if f_obj.get("type") == "timeShortLivedLimit":
                    loiter_time = str(f_obj.get("time", ""))
                    break
                    
        duration_val = loiter_time if loiter_time else "0"
        classes = [oc.get("type","") for oc in scenario.get("objectClassifications",[]) if oc.get("type")]
        target_detection = ", ".join(c.capitalize() + " Detection" for c in classes) if classes else "Any Detection"
        
        rule_type_base = scenario.get("type", "")
        if triggers and not rule_type_base: 
            rule_type_base = triggers[0].get("type", "Unknown")
        rule_type = "Intrusion Detection" if "area" in rule_type_base.lower() else rule_type_base.capitalize()

        # Direct 1:1 Pixel Mapping Engine
        vertices = []
        if triggers:
            raw_vertices = triggers[0].get("vertices", [])
            for pt in raw_vertices:
                raw_x = float(pt[0])
                raw_y = float(pt[1])
                
                # Map Axis normalized [-1.0, 1.0] coordinates to strict pixel dimensions
                # X: -1 (left) to 1 (right)
                pixel_x = int(((raw_x + 1.0) / 2.0) * img_w)
                
                # Y: -1 (bottom) to 1 (top) -> Pillow uses 0 at top, img_h at bottom
                pixel_y = int(((1.0 - raw_y) / 2.0) * img_h)
                
                vertices.append((pixel_x, pixel_y))
                
        parsed_rules.append({
            "is_placeholder": False,
            "name": rule_name,
            "type": rule_type,
            "duration": duration_val,
            "target": target_detection,
            "vertices": vertices
        })
        
    return parsed_rules

def render_axis_overlay(camera_image, vertices, index, img_w, img_h):
    """Draws geometric overlays on the snapshot using Pillow."""
    if camera_image is not None:
        base_img = camera_image.copy().convert("RGBA")
    else:
        base_img = PILImage.new("RGBA", (img_w, img_h), (50, 50, 50, 255))

    # Kingfisher RGB Brand Tuples
    brand_colors = [
        (0, 161, 154),   # Kingfisher Normal
        (0, 114, 110),   # Kingfisher Dark
        (229, 245, 245)  # Kingfisher Light
    ]
    rgb_color = brand_colors[index % 3]
    
    overlay = PILImage.new("RGBA", base_img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    
    if vertices:
        draw.polygon(vertices, fill=rgb_color + (76,))
        draw.polygon(vertices, outline=rgb_color + (255,), width=3)
        
        for (x, y) in vertices:
            r = 6
            draw.ellipse((x - r, y - r, x + r, y + r), fill=(255, 255, 0, 255), outline=(0, 0, 0, 255), width=1)
    
    final_img = PILImage.alpha_composite(base_img, overlay).convert("RGB")
    buf = BytesIO()
    final_img.save(buf, format="PNG")
    buf.seek(0)

    img = PILImage.open(buf)
    ratio = THUMB_H / img.height
    img_resized = img.resize((int(img.width * ratio), THUMB_H), PILImage.Resampling.LANCZOS)
    
    img_buf = BytesIO()
    img_resized.save(img_buf, format="PNG")
    img_buf.seek(0)
    
    return OpenpyxlImage(img_buf)

def create_new_workbook(batch_num):
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Camera Analytics"
    ws_main.views.sheetView[0].showGridLines = True

    ws_missed = wb.create_sheet(title="Missed Cameras")
    ws_missed.views.sheetView[0].showGridLines = True

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="1A1D27")
    failed_fill   = PatternFill("solid", start_color="A61C1C") 
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_main.merge_cells("A1:I1")
    ws_main["A1"] = f"Intelligent Analytics Master Report - Part {batch_num}"
    ws_main["A1"].font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws_main["A1"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A1"].alignment = center_align
    ws_main.row_dimensions[1].height = 36

    ws_main.merge_cells("A2:I2")
    ws_main["A2"] = f"Batch Processing Timeline Context: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_main["A2"].font      = Font(name="Arial", size=10, color="888888")
    ws_main["A2"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A2"].alignment = center_align
    ws_main.row_dimensions[2].height = 20

    main_headers = [
        "Client Name", "Location", "Live Unit Serial", "Camera Position", 
        "Rule Name", "Rule Type", "Target Detection", "Duration (s)", "Rule Visual Overlay Thumbnail"
    ]
    for col, h in enumerate(main_headers, 1):
        cell = ws_main.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        cell.border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                             top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
    ws_main.row_dimensions[3].height = 22

    main_widths = [22, 24, 22, 16, 20, 20, 20, 14, 110]  
    for i, w in enumerate(main_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = w

    ws_missed.merge_cells("A1:G1")
    ws_missed["A1"] = "Analytics Fetch Exception Audit Log"
    ws_missed["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_missed["A1"].fill      = PatternFill("solid", start_color="330000") 
    ws_missed["A1"].alignment = center_align
    ws_missed.row_dimensions[1].height = 36

    missed_headers = ["Timestamp", "Client Name", "Location", "Live Unit Serial", "Camera Port", "Target Endpoint", "Failure Reason"]
    for col, h in enumerate(missed_headers, 1):
        cell = ws_missed.cell(row=2, column=col, value=h)
        cell.font = header_font; cell.fill = failed_fill; cell.alignment = center_align
    ws_missed.row_dimensions[2].height = 22

    missed_widths = [22, 22, 24, 22, 15, 65, 45]
    for i, w in enumerate(missed_widths, 1):
        ws_missed.column_dimensions[get_column_letter(i)].width = w

    ws_main.freeze_panes = "A4"
    ws_missed.freeze_panes = "A3"
    
    return wb, ws_main, ws_missed


# --- 3. DYNAMIC MODE & SELECTION RUNTIME PROMPTS ---
print("=========================================")
print("     AXIS ANALYTICS REPORT GENERATOR     ")
print("=========================================")
print("1. Process a batch of cameras via CSV file")
print("2. Test a single camera / IP address")
print("-----------------------------------------")
run_mode = input("Select an option (1 or 2) [Default 1]: ").strip()

camera_rows = []
base_filename = ""
DATE_SUFFIX = datetime.now().strftime("%Y%m%d")

if run_mode == "2":
    print("\n[ Single Camera Mode Selected ]")
    single_ip = input("Enter the Camera IP Address: ").strip()
    while not single_ip:
        single_ip = input("[!] IP Address cannot be blank. Enter Camera IP: ").strip()
        
    camera_rows.append({
        "CLIENT_NM": input("Enter Client Name (Optional): ").strip() or "Single Test",
        "LOCATION_NM": input("Enter Location Name (Optional): ").strip() or "Diagnostic",
        "LIVE_UNIT_SERIAL_NM": input("Enter Unit Serial (Optional): ").strip() or "N/A",
        "IP": single_ip
    })
    base_filename = f"Single_Test_{single_ip.replace('.', '_')}"
else:
    print("\n[ CSV Batch Mode Selected ]")
    print("Select your input CSV file via the window popup...")
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    selected_file = filedialog.askopenfilename(
        initialdir=downloads_path,
        title="Select Camera Layout CSV File",
        filetypes=(("CSV Files", "*.csv"), ("All Files", "*.*"))
    )
    
    if selected_file:
        csv_input_path = Path(selected_file)
    else:
        print("[-] Window closed without selection. Falling back to manual name lookups.")
        csv_input_path = downloads_path / (input("Enter the input CSV file name: ").strip() or "fullTesting.csv")

    if not csv_input_path.exists():
        print(f"[!] Error: Missing input file profile target: {csv_input_path}")
        exit()

    print(f"Reading target camera layouts from: {csv_input_path}...")
    with open(csv_input_path, mode='r', encoding='utf-8-sig') as f:
        camera_rows = list(csv.DictReader(f))
            
    custom_tag = input("Enter an optional custom tag/job name (or press Enter to use CSV name): ").strip()
    base_filename = custom_tag if custom_tag else csv_input_path.stem

REPORT_PART_FILE_TEMPLATE = f"{base_filename}_Part_{{batch_idx}}_{DATE_SUFFIX}.xlsx"
FINAL_ZIP_ARCHIVE_NAME = f"{base_filename}_Archive_{DATE_SUFFIX}.zip"

# --- 4. EXECUTION ENGINE INIT ---
generated_files = []
batch_index = 1
current_main_row = 4
current_missed_row = 3
rows_written_in_current_file = 0

wb, ws_main, ws_missed = create_new_workbook(batch_index)

normal_font = Font(name="Arial", size=10)
bold_font   = Font(name="Arial", bold=True, size=10)

white_normal_font = Font(name="Arial", size=10, color="FFFFFF")
white_bold_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")

left_align  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
center_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                     top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

# Connection Pool Initialization
session = requests.Session()
auth_strategies = [HTTPDigestAuth(USER, PASSWORD), HTTPBasicAuth(USER, PASSWORD)]

# --- 5. CORE RUNTIME ENGINE LOOP ---
for row_data in camera_rows:
    client_name = row_data.get("CLIENT_NM", "").strip()
    location = row_data.get("LOCATION_NM", "").strip()
    serial = row_data.get("LIVE_UNIT_SERIAL_NM", "").strip()
    ip = row_data.get("IP", "").strip()
    
    if not ip:
        continue

    print(f"\n#################################################################")
    print(f" INGESTING HOST NODE: {client_name} - {location} ({ip})")
    print(f"#################################################################")

    for cam in CAMERA_CONFIGS:
        port = cam["port"]
        position = cam["position"]
        bg_color = cam["color"]
        row_fill = PatternFill("solid", start_color=bg_color)
        
        # Automatically use white text if the background is Kingfisher Dark
        if bg_color == "00726E":  
            active_normal_font = white_normal_font
            active_bold_font   = white_bold_font
        else:
            active_normal_font = normal_font
            active_bold_font   = bold_font

        print(f" -> Testing {position} Interface Port: {port}...")

        # A. FETCH AXIS STREAM BACKGROUND SNAPSHOT
        camera_image = fetch_axis_snapshot(session, ip, port, auth_strategies)
        if camera_image is None:
            print("    [!] Warning: Failed to fetch stream snapshot. Defaulting to 1280x720 canvas.")
            img_w, img_h = 1280, 720
        else:
            img_w, img_h = camera_image.size

        # B. FETCH AXIS OBJECT ANALYTICS CONFIGURATION PAYLOAD
        json_data, control_url, err_msg = fetch_axis_scenarios(session, ip, port, auth_strategies)

        if json_data is None:
            err_msg = err_msg or "Unauthorized Connection (All auth variations failed)"
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws_missed.append([timestamp, client_name, location, serial, f"{position} ({port})", control_url, err_msg])
            for col in range(1, 8):
                ws_missed.cell(row=current_missed_row, column=col).border = thin_border
                ws_missed.cell(row=current_missed_row, column=col).font = normal_font
            current_missed_row += 1
            
            # --- Ensure the snapshot is still logged in the main sheet ---
            rules = [{
                "is_placeholder": True,
                "name": "Analytics Fetch Failed (Check Missed Tab)",
                "type": "N/A",
                "target": "No Analytics Configured",
                "duration": "N/A",
                "vertices": []
            }]
        else:
            # C. PARSE SCENARIOS
            rules = parse_axis_analytics(json_data, img_w, img_h)

        # D. ROW GENERATION BLOCK
        for index, rule in enumerate(rules):
            if rows_written_in_current_file >= MAX_ROWS_PER_FILE:
                file_path = output_dir / REPORT_PART_FILE_TEMPLATE.format(batch_idx=batch_index)
                wb.save(file_path)
                generated_files.append(file_path)
                print(f"    [#] Saved workbook partition part limit reached: {file_path}")
                
                batch_index += 1
                wb, ws_main, ws_missed = create_new_workbook(batch_index)
                current_main_row = 4
                current_missed_row = 3
                rows_written_in_current_file = 0

            def styled_main(col, value, font=active_normal_font, align=left_align):
                c = ws_main.cell(row=current_main_row, column=col, value=value)
                c.font = font; c.fill = row_fill; c.alignment = align; c.border = thin_border
                return c

            styled_main(1, client_name, align=left_align)
            styled_main(2, location, align=left_align)
            styled_main(3, serial, align=center_align)
            styled_main(4, position, font=active_bold_font, align=center_align)
            styled_main(5, rule["name"], align=left_align)
            styled_main(6, rule["type"], align=left_align)
            styled_main(7, rule["target"], align=center_align)
            
            dur = rule["duration"]
            styled_main(8, int(dur) if dur.isdigit() else dur, align=center_align)

            thumb_cell = ws_main.cell(row=current_main_row, column=9, value="")
            thumb_cell.fill = row_fill; thumb_cell.border = thin_border

            # Render Overlay Image with Pillow
            try:
                xl_img = render_axis_overlay(camera_image, rule["vertices"], index, img_w, img_h)
                xl_img.anchor = TwoCellAnchor(editAs="oneCell", 
                                             _from=AnchorMarker(col=8, colOff=0, row=current_main_row-1, rowOff=0), 
                                             to=AnchorMarker(col=9, colOff=0, row=current_main_row, rowOff=0))
                ws_main.add_image(xl_img)
                
                if rule.get("is_placeholder"):
                    print(f"    [+] Logged snapshot row for Port {port} (No rules active or fetch failed)")
                else:
                    print(f"    [+] Logged metrics for Port {port} Scenario: {rule['name']} ({dur}s)")
            except Exception as e:
                thumb_cell.value = f"(Image failed: {e})"

            ws_main.row_dimensions[current_main_row].height = ROW_H
            current_main_row += 1
            rows_written_in_current_file += 1

# --- 6. EXPORT ENGINE ---
file_path = output_dir / REPORT_PART_FILE_TEMPLATE.format(batch_idx=batch_index)

if batch_index == 1 and rows_written_in_current_file < MAX_ROWS_PER_FILE:
    file_path = output_dir / f"{base_filename}_{DATE_SUFFIX}.xlsx"

wb.save(file_path)
generated_files.append(file_path)

# --- 7. CONDITIONAL AUTO-ZIP COMPRESSION MATRIX ENGINE ---
if len(generated_files) > 1:
    zip_output_path = downloads_path / FINAL_ZIP_ARCHIVE_NAME
    print(f"\n[+] Multiple files generated ({len(generated_files)} parts). Packing into a secure ZIP container...")
    
    with zipfile.ZipFile(zip_output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in generated_files:
            zipf.write(file, arcname=file.name)
            os.remove(file)
            
    print(f"[+] Success! Consolidated archives compiled into:\n    {zip_output_path}")
    try:
        os.rmdir(output_dir)
    except Exception:
        pass
else:
    if generated_files:
        print(f"\n[+] Success! All entries processed. Output saved cleanly to:\n    {file_path}")

# --- 8. PERFORMANCE ANALYTICS TIMING BENCHMARK OUTPUT ---
elapsed_seconds = time.time() - script_start_time
minutes = int(elapsed_seconds // 60)
seconds = int(elapsed_seconds % 60)

print(f"\n=================================================================")
print(f" BATCH PROCESSING COMPLETED SUCCESSFULLY")
print(f"=================================================================")
print(f" -> Total Run Duration: {minutes}m {seconds}s ({round(elapsed_seconds, 2)} total seconds)")
print(f" -> Files Output Group: {len(generated_files)} Workbook Sheet Chunk(s)")
print(f"=================================================================\n")