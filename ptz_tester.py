#!/usr/bin/env python3
"""PTZ drift audit -- read-only Excel reporter.

For each unit IP this probes all three camera ports (5010/5015/5020), queries the
camera's LIVE pan/tilt/zoom over the vendor's read-only status endpoint
(camera_engine.get_ptz_endpoint_catalog: Axis ptz.cgi?query=position, Hikvision
/ISAPI/PTZCtrl/channels/1/status), compares it against the expected baseline for
that position, flags drift beyond tolerance, and exports an .xlsx report with a
snapshot per camera. Nothing is ever written to the camera.

Usage:
    python ptz_tester.py <unit-ip> [<unit-ip> ...] [--vendor axis|hikvision]

Credentials come from AXIS_USER/AXIS_PASSWORD (or HIK_*) env vars or access.env --
the same lookup the analytics writer GUI uses. Never hardcode them here: this file
is committed, access.env is not.

Expected baselines are site-specific: add them to EXPECTED_BASELINES below for the
units you audit. A port with no baseline is reported as NO_BASELINE (actual position
recorded, no drift verdict) rather than compared against a made-up zero.
"""

import argparse
import io
import logging
import os
import re
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.auth import HTTPBasicAuth, HTTPDigestAuth

from camera_engine import (
    AxisHandler,
    HikvisionHandler,
    render_overlay_image,
    get_ptz_endpoint_catalog,
    CAMERA_CONFIGS,
    STRICT_TIMEOUT,
    ROW_H,
)

requests.packages.urllib3.disable_warnings()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("PTZExcelAnalyzer")

# =====================================================================
# DRIFT TOLERANCES (READ-ONLY REPORTING THRESHOLDS)
# =====================================================================
TOLERANCES = {
    "minor": {
        "pan": 30.0,   # must exceed 30 deg shortest angular path to flag
        "tilt": 10.0,
    },
    "critical": {
        "pan": 60.0,
        "tilt": 20.0,
    },
}

# Keywords indicating a human operator was actively controlling the camera. Only
# consulted when a log feed is passed to evaluate_unsolicited_drift -- camera log
# retrieval is not wired up yet, so today every out-of-tolerance move is reported
# as unsolicited.
MANUAL_CONTROL_KEYWORDS = [
    "manual", "ptz_control", "user_action", "operator",
    "joystick", "web_ui_move", "ptzctrl", "gui_goto",
]

# Site-specific expected PTZ baselines, keyed by unit IP then camera position.
# Example:
#     "10.0.0.1": {
#         "CENTER": {"pan": 180.0, "tilt": 15.0, "zoom": 1.0},
#         "LEFT":   {"pan": 90.0,  "tilt": 10.0, "zoom": 1.0},
#         "RIGHT":  {"pan": 270.0, "tilt": 10.0, "zoom": 1.0},
#     },
# Zoom is in vendor-native units (Axis 1..9999 lens position, Hik zoom ratio) and is
# reported as an offset only -- it never affects the drift verdict.
EXPECTED_BASELINES: Dict[str, Dict[str, Dict[str, float]]] = {}


# =====================================================================
# 1. ANGULAR WRAP HELPERS
# =====================================================================

def normalize_angle_degrees(val: Optional[float]) -> Optional[float]:
    if val is None:
        return None
    return round(float(val) % 360.0, 2)


def calculate_shortest_angular_distance(actual: float, expected: float) -> Tuple[float, float]:
    """Signed and absolute shortest distance on a 360-degree circle."""
    diff = (actual - expected + 180.0) % 360.0 - 180.0
    return round(diff, 2), round(abs(diff), 2)


def _fmt_offset(val: Any, unit: str = "°") -> str:
    """'+12.5deg' for numbers, 'N/A' for anything else (never crashes on 'N/A')."""
    return f"{val:+}{unit}" if isinstance(val, (int, float)) else "N/A"


# =====================================================================
# 2. CREDENTIALS (env / access.env -- never hardcoded)
# =====================================================================

def load_credentials(vendor: str) -> Tuple[str, str]:
    """Same lookup order as analytics_writer_gui._prefill_credentials."""
    prefix = "HIK" if "hik" in vendor.lower() else "AXIS"
    user = os.environ.get(f"{prefix}_USER")
    passw = os.environ.get(f"{prefix}_PASSWORD") or os.environ.get(f"{prefix}_PASS")
    env_file = Path(__file__).with_name("access.env")
    if (not user or not passw) and env_file.exists():
        # Same corruption guards as the GUI: strip NULs from a UTF-16 append, and
        # truncate a value that has another VAR= assignment fused onto it.
        text = env_file.read_text(errors="ignore").replace("\0", "")
        env = {}
        for k, v in re.findall(r"^([A-Z_]+)=(.*)$", text, re.M):
            glued = re.search(r"[A-Z][A-Z_]{2,}=", v)
            env[k] = (v[:glued.start()] if glued else v).strip()
        user = user or env.get(f"{prefix}_USER")
        passw = passw or env.get(f"{prefix}_PASSWORD") or env.get(f"{prefix}_PASS")
    if not (user and passw):
        raise SystemExit(f"No {prefix}_USER/{prefix}_PASSWORD found in the environment "
                         f"or access.env -- refusing to guess camera credentials.")
    return user, passw


# =====================================================================
# 3. READ-ONLY LIVE QUERIES (POSITION + SNAPSHOT)
# =====================================================================

def fetch_ptz_position(session: requests.Session, ip: str, port: str, vendor: str,
                       user: str, passw: str) -> Dict[str, Optional[float]]:
    """Query the camera's current pan/tilt/zoom. Read-only GET; returns None values
    for anything the camera doesn't report (fixed-lens units, missing endpoint)."""
    catalog = get_ptz_endpoint_catalog(ip, port)
    auths = [HTTPDigestAuth(user, passw), HTTPBasicAuth(user, passw)]
    none = {"pan": None, "tilt": None, "zoom": None}

    if "axis" in vendor.lower():
        url = catalog["axis"]["query_position"]
        for auth in auths:
            try:
                r = session.get(url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
            except requests.exceptions.RequestException:
                continue
            if r.status_code != 200:
                continue
            # Body is plain text lines: pan=179.99 / tilt=0.00 / zoom=333 ...
            vals = {k: float(v) for k, v in re.findall(r"^(\w+)=(-?[\d.]+)\s*$", r.text, re.M)}
            if not vals:
                return none  # answered, but no PTZ readout on this port
            return {"pan": vals.get("pan"), "tilt": vals.get("tilt"), "zoom": vals.get("zoom")}
        return none

    url = catalog["hikvision"]["ptz_status"]
    for auth in auths:
        try:
            r = session.get(url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
        except requests.exceptions.RequestException:
            continue
        if r.status_code != 200:
            continue
        try:
            root = ET.fromstring(r.text)
        except ET.ParseError:
            return none

        def num(tag):
            el = root.find(f".//{{*}}{tag}")
            try:
                return float(el.text) if el is not None and el.text else None
            except ValueError:
                return None

        az, elev, zoom = num("azimuth"), num("elevation"), num("absoluteZoom")
        # ISAPI reports azimuth/elevation/zoom as value x10.
        return {"pan": az / 10.0 if az is not None else None,
                "tilt": elev / 10.0 if elev is not None else None,
                "zoom": zoom / 10.0 if zoom is not None else None}
    return none


def fetch_snapshot_from_engine(
    ip: str, port: str, vendor: str, user: str, passw: str, label: str
) -> Optional[io.BytesIO]:
    """Snapshot via the audit engine's handlers (HTTP GET only, no write commands)."""
    session = requests.Session()
    handler = (AxisHandler(ip, user, passw) if "axis" in vendor.lower()
               else HikvisionHandler(ip, user, passw))

    camera_image = None
    try:
        camera_image, snap_url, snap_err, auth_rejected = handler.fetch_snapshot(session, port)
        if camera_image is not None:
            img_w, img_h = camera_image.size
            logger.info(f"Successfully fetched snapshot from {ip}:{port} ({label})")
        else:
            img_w, img_h = handler.fallback_dim
            logger.warning(f"Snapshot HTTP fetch failed for {ip}:{port} ({snap_err}). Using fallback frame.")
    except Exception as exc:
        img_w, img_h = handler.fallback_dim
        logger.warning(f"Error connecting to {ip}:{port} ({exc}). Using fallback frame.")

    try:
        img_buf = render_overlay_image(
            camera_image=camera_image,
            vertices=[],
            index=0,
            img_w=img_w,
            img_h=img_h,
            rule_type=f"PTZ Audit: {label}",
        )
        if camera_image is not None:
            camera_image.close()
        if img_buf:
            img_buf.seek(0)  # openpyxl reads from byte 0
        return img_buf
    except Exception as exc:
        logger.error(f"Failed to render overlay buffer for {ip}:{port}: {exc}")
        return None


# =====================================================================
# 4. DRIFT ANALYSIS & TIMESTAMP PARSING
# =====================================================================

def parse_log_timestamp(log_entry: str) -> str:
    """Best-effort ISO or syslog timestamp out of a camera log line."""
    iso_match = re.search(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}", log_entry)
    if iso_match:
        return iso_match.group(0)
    syslog_match = re.search(r"[A-Z][a-z]{2}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}", log_entry)
    if syslog_match:
        return syslog_match.group(0)
    return "TIMESTAMP_UNKNOWN"


def evaluate_unsolicited_drift(
    actual: Dict[str, Optional[float]],
    expected: Dict[str, float],
    logs: List[str],
) -> Dict[str, Any]:
    metrics: Dict[str, Any] = {}
    severity_score = 0
    directions = []

    # Newest log first: was an operator actively moving the camera?
    is_manual_action_logged = False
    operator_timestamp = "N/A"
    for log_entry in reversed(logs):
        if any(kw in log_entry.lower() for kw in MANUAL_CONTROL_KEYWORDS):
            is_manual_action_logged = True
            operator_timestamp = parse_log_timestamp(log_entry)
            break

    # 1. Pan (circular math)
    act_pan = normalize_angle_degrees(actual.get("pan"))
    exp_pan = normalize_angle_degrees(expected.get("pan"))
    if act_pan is not None and exp_pan is not None:
        raw_pan_diff, abs_pan_diff = calculate_shortest_angular_distance(act_pan, exp_pan)
        metrics.update(pan_actual=act_pan, pan_expected=exp_pan,
                       pan_raw_diff=raw_pan_diff, pan_abs_diff=abs_pan_diff)
        if abs_pan_diff >= TOLERANCES["minor"]["pan"]:
            directions.append("RIGHT" if raw_pan_diff > 0 else "LEFT")
            severity_score += 4 if abs_pan_diff >= TOLERANCES["critical"]["pan"] else 1
    else:
        metrics.update(pan_actual=act_pan if act_pan is not None else "N/A",
                       pan_expected=exp_pan if exp_pan is not None else "N/A",
                       pan_raw_diff="N/A")

    # 2. Tilt (linear)
    act_tilt = actual.get("tilt")
    exp_tilt = expected.get("tilt")
    if act_tilt is not None and exp_tilt is not None:
        raw_tilt_diff = round(float(act_tilt) - float(exp_tilt), 2)
        abs_tilt_diff = abs(raw_tilt_diff)
        metrics.update(tilt_actual=round(float(act_tilt), 2), tilt_expected=exp_tilt,
                       tilt_raw_diff=raw_tilt_diff, tilt_abs_diff=abs_tilt_diff)
        if abs_tilt_diff >= TOLERANCES["minor"]["tilt"]:
            directions.append("UP" if raw_tilt_diff > 0 else "DOWN")
            severity_score += 4 if abs_tilt_diff >= TOLERANCES["critical"]["tilt"] else 1
    else:
        metrics.update(tilt_actual=act_tilt if act_tilt is not None else "N/A",
                       tilt_expected=exp_tilt if exp_tilt is not None else "N/A",
                       tilt_raw_diff="N/A")

    # 3. Zoom (vendor-native units; reported only, never drives the verdict)
    act_zoom, exp_zoom = actual.get("zoom"), expected.get("zoom")
    if act_zoom is not None and exp_zoom is not None:
        metrics.update(zoom_actual=act_zoom, zoom_expected=exp_zoom,
                       zoom_raw_diff=round(float(act_zoom) - float(exp_zoom), 2))
    else:
        metrics.update(zoom_actual=act_zoom if act_zoom is not None else "N/A",
                       zoom_expected=exp_zoom if exp_zoom is not None else "N/A",
                       zoom_raw_diff="N/A")

    # Verdict
    if actual.get("pan") is None and actual.get("tilt") is None:
        status = "NO_PTZ_DATA"          # camera never reported a position
    elif not expected:
        status = "NO_BASELINE"          # nothing to compare against for this port
    elif severity_score == 0:
        status = "OK"
    elif is_manual_action_logged:
        status = "MANUAL_OVERRIDE_OK"
    elif severity_score >= 4:
        status = "CRITICAL_UNSOLICITED_DRIFT"
    else:
        status = "WARNING_UNSOLICITED_DRIFT"

    metrics["status"] = status
    metrics["severity_score"] = severity_score
    metrics["drift_direction"] = " | ".join(directions) if directions else "CENTERED"
    metrics["is_manual_action"] = is_manual_action_logged
    metrics["operator_timestamp"] = operator_timestamp
    return metrics


# =====================================================================
# 5. EXCEL REPORT EXPORTER WITH EMBEDDED SNAPSHOTS
# =====================================================================

STATUS_COLORS = {
    "CRITICAL_UNSOLICITED_DRIFT": "FFC7CE",  # soft red
    "WARNING_UNSOLICITED_DRIFT": "FFEB9C",   # soft yellow
    "MANUAL_OVERRIDE_OK": "D9EAD3",          # soft sage green
    "NO_PTZ_DATA": "EEEEEE",                 # grey -- no position readout
    "NO_BASELINE": "DDEBF7",                 # light blue -- recorded, not judged
}


def export_reports_to_excel(reports: List[Dict[str, Any]], output_file: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PTZ Drift Audit"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1A1D27")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD")
    )

    ws.merge_cells("A1:N1")
    ws["A1"] = "Multi-Camera PTZ Drift Audit Master Report (Read-Only)"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="0F1117")
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 36

    headers = [
        "Unit ID", "Camera Position", "Vendor", "IP : Port",
        "Overall Status", "Manual Override?", "Operator Move Time", "Direction Vector",
        "Pan Offset", "Tilt Offset", "Zoom Offset", "Severity Score",
        "Scan Timestamp", "Camera Snapshot View"
    ]
    widths = [16, 16, 12, 18, 28, 22, 24, 20, 14, 14, 14, 15, 22, 95]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    ws.row_dimensions[2].height = 25

    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A3"

    current_row = 3
    for report in reports:
        status = report.get("overall_status", "OK")
        bg_color = STATUS_COLORS.get(status, report.get("bg_color", "E5F5F5"))
        row_fill = PatternFill("solid", start_color=bg_color)

        data_row = [
            report.get("unit_id"),
            report.get("position"),
            report.get("vendor", "").upper(),
            report.get("ip_address"),
            report.get("overall_status"),
            "YES (Operator Moved)" if report.get("is_manual_action") else "NO (Unsolicited)",
            report.get("operator_timestamp"),
            report.get("drift_direction_vector"),
            _fmt_offset(report.get("pan_offset")),
            _fmt_offset(report.get("tilt_offset")),
            _fmt_offset(report.get("zoom_offset"), unit=""),
            report.get("severity_score"),
            report.get("timestamp_utc"),
        ]

        for col_idx, val in enumerate(data_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill = row_fill; cell.alignment = center_align; cell.border = thin_border; cell.font = Font(name="Arial", size=10)

        thumb_cell = ws.cell(row=current_row, column=14, value="")
        thumb_cell.fill = row_fill; thumb_cell.border = thin_border

        img_buf = report.get("snapshot_buffer")
        if img_buf:
            img_buf.seek(0)
            xl_img = OpenpyxlImage(img_buf)
            # Anchor inside column N (0-index 13) so images don't drift/overlap.
            xl_img.anchor = TwoCellAnchor(
                editAs="oneCell",
                _from=AnchorMarker(col=13, colOff=0, row=current_row - 1, rowOff=0),
                to=AnchorMarker(col=13, colOff=0, row=current_row, rowOff=0),
            )
            ws.add_image(xl_img)

        ws.row_dimensions[current_row].height = ROW_H
        current_row += 1

    output_path = Path(output_file)
    wb.save(output_path)
    print(f"\n[+] Read-Only Excel Report saved to: {output_path.resolve()}\n")


# =====================================================================
# 6. MAIN SCANNER RUNNER
# =====================================================================

STATUS_BADGES = {
    "CRITICAL_UNSOLICITED_DRIFT": "[CRITICAL UNSOLICITED DRIFT]",
    "WARNING_UNSOLICITED_DRIFT": "[WARNING UNSOLICITED DRIFT]",
    "MANUAL_OVERRIDE_OK": "[MANUAL OPERATOR ACTION (IGNORED)]",
    "NO_PTZ_DATA": "[NO PTZ READOUT]",
    "NO_BASELINE": "[NO BASELINE CONFIGURED]",
    "OK": "[OK]",
}


def run_multi_camera_ptz_scan(units: List[Dict[str, Any]],
                              output_excel: str = "PTZ_Multi_Camera_Report.xlsx"):
    all_reports = []

    print("\n" + "=" * 75)
    print("      MULTI-CAMERA LOCAL PTZ DRIFT & VIEW ANALYSIS SCANNER      ")
    print("=" * 75 + "\n")

    for unit in units:
        unit_id = unit["unit_id"]
        ip = unit["ip"]
        vendor = unit["vendor"].lower()
        user = unit["username"]
        pwd = unit["password"]
        baselines = unit.get("expected_baselines", {})
        session = requests.Session()

        print(f"Auditing Unit [{unit_id}] at Base IP: {ip} (Read-Only)")
        print("-" * 75)

        for cam in CAMERA_CONFIGS:
            port = cam["port"]
            position = cam["position"]
            expected = baselines.get(position, {})

            actual_ptz = fetch_ptz_position(session, ip, port, vendor, user, pwd)
            # Camera-log retrieval (operator moves with timestamps) is not wired up
            # yet, so no MANUAL_OVERRIDE_OK verdicts are possible today -- every
            # out-of-tolerance move reports as unsolicited until a log feed exists.
            metrics = evaluate_unsolicited_drift(actual_ptz, expected, logs=[])
            snapshot_buffer = fetch_snapshot_from_engine(ip, port, vendor, user, pwd,
                                                         f"{unit_id} - {position}")

            all_reports.append({
                "timestamp_utc": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "unit_id": unit_id,
                "position": position,
                "vendor": vendor,
                "ip_address": f"{ip}:{port}",
                "bg_color": cam["color"],
                "overall_status": metrics["status"],
                "is_manual_action": metrics["is_manual_action"],
                "operator_timestamp": metrics["operator_timestamp"],
                "severity_score": metrics["severity_score"],
                "drift_direction_vector": metrics["drift_direction"],
                "pan_actual": metrics["pan_actual"],
                "pan_expected": metrics["pan_expected"],
                "pan_offset": metrics["pan_raw_diff"],
                "tilt_actual": metrics["tilt_actual"],
                "tilt_expected": metrics["tilt_expected"],
                "tilt_offset": metrics["tilt_raw_diff"],
                "zoom_actual": metrics["zoom_actual"],
                "zoom_expected": metrics["zoom_expected"],
                "zoom_offset": metrics["zoom_raw_diff"],
                "snapshot_buffer": snapshot_buffer,
            })

            print(f"  --> Camera Position : {position:<8} (Port {port})")
            print(f"      Status          : {STATUS_BADGES.get(metrics['status'], metrics['status'])}")
            print(f"      Live PTZ        : pan={metrics['pan_actual']} tilt={metrics['tilt_actual']} "
                  f"zoom={metrics['zoom_actual']}")
            print(f"      Pan Offset      : {_fmt_offset(metrics['pan_raw_diff'])} | "
                  f"Tilt Offset: {_fmt_offset(metrics['tilt_raw_diff'])}")
            print(f"      Snapshot State  : {'Loaded into memory' if snapshot_buffer else 'Fallback generated'}")
            print()

        print("-" * 75 + "\n")

    export_reports_to_excel(all_reports, output_excel)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Read-only PTZ drift audit: live position + snapshot per port, "
                    "Excel report out. Baselines live in EXPECTED_BASELINES.")
    ap.add_argument("ips", nargs="+", help="unit base IP(s); each probes ports 5010/5015/5020")
    ap.add_argument("--vendor", choices=["axis", "hikvision"], default="axis",
                    help="camera vendor for all listed units (default: axis)")
    ap.add_argument("--output", default="PTZ_Multi_Camera_Report.xlsx",
                    help="output .xlsx path")
    args = ap.parse_args()

    user, passw = load_credentials(args.vendor)
    units = [{
        "unit_id": ip,
        "vendor": args.vendor,
        "ip": ip,
        "username": user,
        "password": passw,
        "expected_baselines": EXPECTED_BASELINES.get(ip, {}),
    } for ip in args.ips]
    run_multi_camera_ptz_scan(units, args.output)
