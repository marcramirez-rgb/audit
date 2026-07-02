import os
import csv
import json
import time 
import urllib3
import zipfile
import xml.etree.ElementTree as ET
import platform
import subprocess
from io import BytesIO
from pathlib import Path
from datetime import datetime
import requests
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
from PIL import Image as PILImage, ImageDraw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

# --- GUI & Environment Imports ---
import tkinter as tk
from tkinter import filedialog
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Start the global execution benchmark timer
script_start_time = time.time()

# Suppress self-signed certificate warnings 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- 1. SET UP INFRASTRUCTURE PATHING & CREDENTIALS ---
USER_HIK = os.getenv("HIK_USER", "admin")
PASS_HIK = os.getenv("HIK_PASSWORD")
USER_AXIS = os.getenv("AXIS_USER", "root")
PASS_AXIS = os.getenv("AXIS_PASSWORD")

STRICT_TIMEOUT = (3.05, 5.0)
THUMB_H = 360  
ROW_H   = 275  
MAX_ROWS_PER_FILE = 100 

downloads_path = Path.home() / "Downloads"
output_dir = downloads_path / "Camera_Reports_Batch"
output_dir.mkdir(parents=True, exist_ok=True)

# Lens-specific port configurations (Using Kingfisher Brand Hex Codes)
CAMERA_CONFIGS = [
    {"port": "5010", "position": "CENTER", "color": "E5F5F5"},  # Kingfisher Light
    {"port": "5015", "position": "LEFT",   "color": "00A19A"},  # Kingfisher Normal
    {"port": "5020", "position": "RIGHT",  "color": "00726E"}   # Kingfisher Dark
]


# --- 2. HIKVISION SPECIFIC WORKERS ---

def fetch_hik_snapshot(session, ip, port, auth_strategies):
    img_url = f"http://{ip}:{port}/ISAPI/Streaming/channels/101/picture"
    for auth in auth_strategies:
        try:
            with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, stream=True, verify=False) as response:
                if response.status_code == 200:
                    return PILImage.open(BytesIO(response.content))
        except requests.exceptions.RequestException:
            continue
    return None

def fetch_hik_rules(session, ip, port, auth_strategies):
    for channel_id in [1, 2]:
        rule_url = f"http://{ip}:{port}/ISAPI/Intelligent/channels/{channel_id}/behaviorRule/1"
        for auth in auth_strategies:
            try:
                response = session.get(rule_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
                if response.status_code == 200:
                    temp_xml = response.text
                    if temp_xml and ("positionX" in temp_xml or "RegionCoordinates" in temp_xml):
                        return temp_xml, rule_url, None
            except requests.exceptions.RequestException:
                continue
    return None, f"http://{ip}:{port}/ISAPI/Intelligent/channels/[1,2]/behaviorRule/1", "No active analytic perimeters found"

def parse_hik_analytics(xml_data, img_w, img_h):
    namespaces = {'ns': 'http://www.std-cgi.com/ver20/XMLSchema'}
    parsed_rules = []
    
    try:
        root = ET.fromstring(xml_data)
        rules = root.findall('.//ns:RuleInfo', namespaces)
    except Exception:
        rules = []

    if not rules:
        return [{
            "is_placeholder": True, "name": "No Scenarios Configured", 
            "type": "N/A", "target": "No Analytics Configured", 
            "duration": "N/A", "vertices": []
        }]

    for rule in rules:
        rule_name = rule.find('ns:ruleName', namespaces).text
        
        event_type_raw = rule.find('ns:eventType', namespaces)
        event_type = event_type_raw.text if event_type_raw is not None else "Unknown"
        if "field" in event_type.lower() or "intrusion" in event_type.lower():
            event_type = "Intrusion Detection"
        elif "line" in event_type.lower() or "cross" in event_type.lower():
            event_type = "Line Crossing"

        duration_val = "0"
        possible_tags = [
            './/ns:FieldDetectionParam/ns:durationTime', './/ns:FieldDetectionParam/ns:timeDuration',
            './/ns:LineCrossing/ns:duration', './/ns:Intrusion/ns:duration',
            './/ns:durationTime', './/ns:timeDuration', './/ns:duration', './/ns:alarmDelay'
        ]
        found_values = [rule.find(tag, namespaces).text.strip() for tag in possible_tags if rule.find(tag, namespaces) is not None and rule.find(tag, namespaces).text]
        if found_values:
            valid_seconds = [v for v in found_values if v != "100" and v != "0"]
            duration_val = valid_seconds[0] if valid_seconds else found_values[0]

        target_node = rule.find('.//ns:FieldDetectionParam/ns:detectionTarget', namespaces)
        if target_node is None:
            target_node = rule.find('.//ns:detectionTarget', namespaces) or rule.find('.//ns:TargetType', namespaces)
        target_detection = target_node.text.capitalize() if (target_node is not None and target_node.text) else "All Targets"

        region = rule.find('.//ns:RegionCoordinatesList', namespaces)
        if region is None:
            continue
            
        vertices = []
        for coord in region.findall('ns:RegionCoordinates', namespaces):
            raw_x = float(coord.find('ns:positionX', namespaces).text)
            raw_y = float(coord.find('ns:positionY', namespaces).text)
            
            pixel_x = int((raw_x / 1000.0) * img_w)
            pixel_y = int((raw_y / 1000.0) * img_h)
            
            pixel_x = img_w - pixel_x
            pixel_y = img_h - pixel_y
            
            vertices.append((pixel_x, pixel_y))
            
        if not vertices:
            continue
            
        parsed_rules.append({
            "is_placeholder": False,
            "name": rule_name,
            "type": event_type,
            "duration": duration_val,
            "target": target_detection,
            "vertices": vertices
        })
        
    return parsed_rules if parsed_rules else [{
        "is_placeholder": True, "name": "No Scenarios Configured", 
        "type": "N/A", "target": "No Analytics Configured", 
        "duration": "N/A", "vertices": []
    }]


# --- 3. AXIS SPECIFIC WORKERS ---

def fetch_axis_snapshot(session, ip, port, auth_strategies):
    img_url = f"http://{ip}:{port}/axis-cgi/jpg/image.cgi?resolution=1280x720"
    for auth in auth_strategies:
        try:
            with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, stream=True, verify=False) as response:
                if response.status_code == 200:
                    return PILImage.open(BytesIO(response.content))
        except requests.exceptions.RequestException:
            continue
    return None

def fetch_axis_scenarios(session, ip, port, auth_strategies):
    control_url = f"http://{ip}:{port}/local/objectanalytics/control.cgi"
    payload = {"apiVersion": "1.2", "method": "getConfiguration"}
    
    for auth in auth_strategies:
        try:
            response = session.post(control_url, auth=auth, json=payload, timeout=STRICT_TIMEOUT, verify=False)
            if response.status_code == 200:
                return response.json(), control_url, None
            else:
                last_err = f"HTTP Error Status {response.status_code}"
        except Exception as e:
            last_err = str(e)
            continue
            
    return None, control_url, last_err

def parse_axis_analytics(json_data, img_w, img_h):
    scenarios = json_data.get("data", {}).get("scenarios", [])
    if not scenarios:
        return [{
            "is_placeholder": True, "name": "No Scenarios Configured", 
            "type": "N/A", "target": "No Analytics Configured", 
            "duration": "N/A", "vertices": []
        }]
        
    parsed_rules = []
    for scenario in scenarios:
        if scenario.get("is_placeholder"):
            parsed_rules.append({
                "is_placeholder": True, "name": "No Scenarios Configured", 
                "type": "N/A", "target": "No Analytics Configured", 
                "duration": "N/A", "vertices": []
            })
            continue

        rule_name = scenario.get("name", f"Scenario {scenario.get('id')}")
        loiter_time = ""
        triggers = scenario.get("triggers", [])
        
        for trigger_obj in triggers:
            for condition in trigger_obj.get("conditions", []):
                if condition.get("type") == "individualTimeInArea":
                    for d in condition.get("data", []):
                        if d.get("time"):
                            loiter_time = str(d["time"])
                            break
                            
        if not loiter_time:
            for f_obj in scenario.get("filters", []):
                if f_obj.get("type") == "timeShortLivedLimit":
                    loiter_time = str(f_obj.get("time", ""))
                    break
                    
        duration_val = loiter_time if loiter_time else "0"
        classes = [oc.get("type","") for oc in scenario.get("objectClassifications",[]) if oc.get("type")]
        target_detection = ", ".join(c.capitalize() + " Detection" for c in classes) if classes else "Any Detection"
        
        rule_type_base = scenario.get("type", "")
        if triggers and not rule_type_base: 
            rule_type_base = triggers[0].get("type", "Unknown")
        rule_type = "Intrusion Detection" if "area" in rule_type_base.lower() else rule_type_base.capitalize()

        vertices = []
        if triggers:
            raw_vertices = triggers[0].get("vertices", [])
            for pt in raw_vertices:
                raw_x = float(pt[0])
                raw_y = float(pt[1])
                pixel_x = int(((raw_x + 1.0) / 2.0) * img_w)
                pixel_y = int(((1.0 - raw_y) / 2.0) * img_h)
                vertices.append((pixel_x, pixel_y))
                
        parsed_rules.append({
            "is_placeholder": False, "name": rule_name, "type": rule_type,
            "duration": duration_val, "target": target_detection, "vertices": vertices
        })
        
    return parsed_rules


# --- 4. UNIVERSAL RENDERING & EXCEL ENGINE ---

def render_overlay_image(camera_image, vertices, index, img_w, img_h):
    if camera_image is not None:
        base_img = camera_image.copy().convert("RGBA")
    else:
        base_img = PILImage.new("RGBA", (img_w, img_h), (50, 50, 50, 255))

    brand_colors = [(0, 161, 154), (0, 114, 110), (229, 245, 245)]
    rgb_color = brand_colors[index % 3]
    
    overlay = PILImage.new("RGBA", base_img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    
    if vertices:
        draw.polygon(vertices, fill=rgb_color + (76,))
        draw.polygon(vertices, outline=rgb_color + (255,), width=3)
        for (x, y) in vertices:
            r = 6
            draw.ellipse((x - r, y - r, x + r, y + r), fill=(255, 255, 0, 255), outline=(0, 0, 0, 255), width=1)
    
    final_img = PILImage.alpha_composite(base_img, overlay).convert("RGB")
    buf = BytesIO()
    final_img.save(buf, format="PNG")
    buf.seek(0)

    img = PILImage.open(buf)
    ratio = THUMB_H / img.height
    img_resized = img.resize((int(img.width * ratio), THUMB_H), PILImage.Resampling.LANCZOS)
    
    img_buf = BytesIO()
    img_resized.save(img_buf, format="PNG")
    img_buf.seek(0)
    
    return OpenpyxlImage(img_buf)

def create_new_workbook(batch_num):
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Camera Analytics"
    ws_main.views.sheetView[0].showGridLines = True

    ws_missed = wb.create_sheet(title="Missed Cameras")
    ws_missed.views.sheetView[0].showGridLines = True

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="1A1D27")
    failed_fill   = PatternFill("solid", start_color="A61C1C") 
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_main.merge_cells("A1:I1")
    ws_main["A1"] = f"Intelligent Analytics Master Report - Part {batch_num}"
    ws_main["A1"].font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws_main["A1"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A1"].alignment = center_align
    ws_main.row_dimensions[1].height = 36

    ws_main.merge_cells("A2:I2")
    ws_main["A2"] = f"Batch Processing Timeline Context: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_main["A2"].font      = Font(name="Arial", size=10, color="888888")
    ws_main["A2"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A2"].alignment = center_align
    ws_main.row_dimensions[2].height = 20

    main_headers = [
        "Client Name", "Location", "Live Unit Serial", "Camera Position", 
        "Rule Name", "Rule Type", "Target Detection", "Duration (s)", "Rule Visual Overlay Thumbnail"
    ]
    for col, h in enumerate(main_headers, 1):
        cell = ws_main.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        cell.border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                             top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
    ws_main.row_dimensions[3].height = 22

    main_widths = [22, 24, 22, 16, 20, 20, 20, 14, 110]  
    for i, w in enumerate(main_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = w

    ws_missed.merge_cells("A1:G1")
    ws_missed["A1"] = "Analytics Fetch Exception Audit Log"
    ws_missed["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_missed["A1"].fill      = PatternFill("solid", start_color="330000") 
    ws_missed["A1"].alignment = center_align
    ws_missed.row_dimensions[1].height = 36

    missed_headers = ["Timestamp", "Client Name", "Location", "Live Unit Serial", "Camera Port", "Target Endpoint", "Failure Reason"]
    for col, h in enumerate(missed_headers, 1):
        cell = ws_missed.cell(row=2, column=col, value=h)
        cell.font = header_font; cell.fill = failed_fill; cell.alignment = center_align
    ws_missed.row_dimensions[2].height = 22

    missed_widths = [22, 22, 24, 22, 15, 65, 45]
    for i, w in enumerate(missed_widths, 1):
        ws_missed.column_dimensions[get_column_letter(i)].width = w

    ws_main.freeze_panes = "A4"
    ws_missed.freeze_panes = "A3"
    
    return wb, ws_main, ws_missed


# --- 5. DYNAMIC MODE & SELECTION RUNTIME PROMPTS ---
print("=========================================")
print("  UNIFIED ANALYTICS REPORT GENERATOR     ")
print("=========================================")
print("1. Process a batch of cameras via CSV file")
print("2. Test a single camera / IP address")
print("-----------------------------------------")
run_mode = input("Select an option (1 or 2) [Default 1]: ").strip()

camera_rows = []
base_filename = ""
DATE_SUFFIX = datetime.now().strftime("%Y%m%d")

if run_mode == "2":
    print("\n[ Single Camera Mode Selected ]")
    single_ip = input("Enter the Camera IP Address: ").strip()
    while not single_ip:
        single_ip = input("[!] IP Address cannot be blank. Enter Camera IP: ").strip()
        
    print("\nSelect Camera Manufacturer:")
    print("1. Hikvision")
    print("2. Axis")
    mfg_choice = input("Enter 1 or 2 [Default 1]: ").strip()
    single_mfg = "Axis" if mfg_choice == "2" else "Hikvision"

    camera_rows.append({
        "CLIENT_NM": input("\nEnter Client Name (Optional): ").strip() or "Single Test",
        "LOCATION_NM": input("Enter Location Name (Optional): ").strip() or "Diagnostic",
        "LIVE_UNIT_SERIAL_NM": input("Enter Unit Serial (Optional): ").strip() or "N/A",
        "IP": single_ip,
        "MANUFACTURER": single_mfg
    })
    base_filename = f"Diagnostic_Test_{single_ip.replace('.', '_')}"
else:
    print("\n[ CSV Batch Mode Selected ]")
    print("Select your input CSV file via the window popup...")
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    selected_file = filedialog.askopenfilename(
        initialdir=downloads_path,
        title="Select Camera Layout CSV File",
        filetypes=(("CSV Files", "*.csv"), ("All Files", "*.*"))
    )
    
    if selected_file:
        csv_input_path = Path(selected_file)
    else:
        print("[-] Window closed without selection. Falling back to manual name lookups.")
        csv_input_path = downloads_path / (input("Enter the input CSV file name: ").strip() or "fullTesting.csv")

    if not csv_input_path.exists():
        print(f"[!] Error: Missing input file profile target: {csv_input_path}")
        exit()

    print(f"Reading target camera layouts from: {csv_input_path}...")
    with open(csv_input_path, mode='r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = [col.strip().upper() for col in reader.fieldnames] if reader.fieldnames else []
        
        # PROACTIVE CSV VALIDATION
        if "MANUFACTURER" not in headers:
            ans = input("\n[!] Warning: 'MANUFACTURER' column missing from CSV.\n    Defaulting all cameras to Hikvision. Continue? (Y/N): ").strip().upper()
            if ans != 'Y':
                print("[-] Exiting process. Please update your CSV file.")
                exit()
                
        camera_rows = list(reader)
            
    custom_tag = input("\nEnter an optional custom tag/job name (or press Enter to use CSV name): ").strip()
    base_filename = custom_tag if custom_tag else csv_input_path.stem

REPORT_PART_FILE_TEMPLATE = f"{base_filename}_Part_{{batch_idx}}_{DATE_SUFFIX}.xlsx"
FINAL_ZIP_ARCHIVE_NAME = f"{base_filename}_Archive_{DATE_SUFFIX}.zip"

# --- 6. EXECUTION ENGINE INIT ---
generated_files = []
batch_index = 1
current_main_row = 4
current_missed_row = 3
rows_written_in_current_file = 0

wb, ws_main, ws_missed = create_new_workbook(batch_index)

normal_font = Font(name="Arial", size=10)
bold_font   = Font(name="Arial", bold=True, size=10)
white_normal_font = Font(name="Arial", size=10, color="FFFFFF")
white_bold_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")

left_align  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
center_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                     top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

session = requests.Session()
total_cameras = len(camera_rows)

# --- 7. CORE RUNTIME ENGINE LOOP ---
for row_idx, row_data in enumerate(camera_rows, 1):
    client_name = row_data.get("CLIENT_NM", "").strip()
    location = row_data.get("LOCATION_NM", "").strip()
    serial = row_data.get("LIVE_UNIT_SERIAL_NM", "").strip()
    ip = row_data.get("IP", "").strip()
    
    # Safely handle the Manufacturer column missing entirely
    mfg_str = row_data.get("MANUFACTURER", "")
    if mfg_str is None:
        mfg_str = ""
    mfg_str = mfg_str.strip().lower()
    
    if not ip:
        continue

    is_axis = "axis" in mfg_str
    api_target_str = "AXIS" if is_axis else "HIKVISION"

    print(f"\n#################################################################")
    # BATCH PROGRESS TRACKING ADDED
    print(f" INGESTING HOST NODE [{row_idx}/{total_cameras}]: {client_name} - {location} ({ip}) [{api_target_str}]")
    print(f"#################################################################")

    # JIT (JUST-IN-TIME) PASSWORD PROMPTING
    if is_axis and not PASS_AXIS:
        PASS_AXIS = input("\n[!] AXIS_PASSWORD not found. Please enter Axis camera password: ").strip()
        if not PASS_AXIS:
            print("[-] Error: Password cannot be blank. Skipping camera.")
            continue
    elif not is_axis and not PASS_HIK:
        PASS_HIK = input("\n[!] HIK_PASSWORD not found. Please enter Hikvision camera password: ").strip()
        if not PASS_HIK:
            print("[-] Error: Password cannot be blank. Skipping camera.")
            continue

    active_user = USER_AXIS if is_axis else USER_HIK
    active_pass = PASS_AXIS if is_axis else PASS_HIK
    auth_strategies = [HTTPDigestAuth(active_user, active_pass), HTTPBasicAuth(active_user, active_pass)]

    for cam in CAMERA_CONFIGS:
        port = cam["port"]
        position = cam["position"]
        bg_color = cam["color"]
        row_fill = PatternFill("solid", start_color=bg_color)
        
        if bg_color == "00726E":  
            active_normal_font = white_normal_font
            active_bold_font   = white_bold_font
        else:
            active_normal_font = normal_font
            active_bold_font   = bold_font

        print(f" -> Testing {position} Interface Port: {port}...")

        if is_axis:
            camera_image = fetch_axis_snapshot(session, ip, port, auth_strategies)
            img_w, img_h = (1280, 720) if camera_image is None else camera_image.size
        else:
            camera_image = fetch_hik_snapshot(session, ip, port, auth_strategies)
            img_w, img_h = (1920, 1080) if camera_image is None else camera_image.size

        if camera_image is None:
            print(f"    [!] Warning: Failed to fetch stream snapshot. Defaulting to {img_w}x{img_h} canvas.")

        if is_axis:
            payload_data, req_url, err_msg = fetch_axis_scenarios(session, ip, port, auth_strategies)
        else:
            payload_data, req_url, err_msg = fetch_hik_rules(session, ip, port, auth_strategies)

        if payload_data is None:
            err_msg = err_msg or "Unauthorized Connection (All auth variations failed)"
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws_missed.append([timestamp, client_name, location, serial, f"{position} ({port})", req_url, err_msg])
            for col in range(1, 8):
                ws_missed.cell(row=current_missed_row, column=col).border = thin_border
                ws_missed.cell(row=current_missed_row, column=col).font = normal_font
            current_missed_row += 1
            
            rules = [{
                "is_placeholder": True,
                "name": "Analytics Fetch Failed (Check Missed Tab)",
                "type": "N/A",
                "target": "No Analytics Configured",
                "duration": "N/A",
                "vertices": []
            }]
        else:
            if is_axis:
                rules = parse_axis_analytics(payload_data, img_w, img_h)
            else:
                rules = parse_hik_analytics(payload_data, img_w, img_h)

        for index, rule in enumerate(rules):
            if rows_written_in_current_file >= MAX_ROWS_PER_FILE:
                file_path = output_dir / REPORT_PART_FILE_TEMPLATE.format(batch_idx=batch_index)
                wb.save(file_path)
                generated_files.append(file_path)
                print(f"    [#] Saved workbook partition part limit reached: {file_path}")
                
                batch_index += 1
                wb, ws_main, ws_missed = create_new_workbook(batch_index)
                current_main_row = 4
                current_missed_row = 3
                rows_written_in_current_file = 0

            def styled_main(col, value, font=active_normal_font, align=left_align):
                c = ws_main.cell(row=current_main_row, column=col, value=value)
                c.font = font; c.fill = row_fill; c.alignment = align; c.border = thin_border
                return c

            styled_main(1, client_name, align=left_align)
            styled_main(2, location, align=left_align)
            styled_main(3, serial, align=center_align)
            styled_main(4, position, font=active_bold_font, align=center_align)
            styled_main(5, rule["name"], align=left_align)
            styled_main(6, rule["type"], align=left_align)
            styled_main(7, rule["target"], align=center_align)
            
            dur = rule["duration"]
            styled_main(8, int(dur) if dur.isdigit() else dur, align=center_align)

            thumb_cell = ws_main.cell(row=current_main_row, column=9, value="")
            thumb_cell.fill = row_fill; thumb_cell.border = thin_border

            try:
                xl_img = render_overlay_image(camera_image, rule["vertices"], index, img_w, img_h)
                xl_img.anchor = TwoCellAnchor(editAs="oneCell", 
                                             _from=AnchorMarker(col=8, colOff=0, row=current_main_row-1, rowOff=0), 
                                             to=AnchorMarker(col=9, colOff=0, row=current_main_row, rowOff=0))
                ws_main.add_image(xl_img)
                
                if rule.get("is_placeholder"):
                    print(f"    [+] Logged snapshot row for Port {port} (No rules active or fetch failed)")
                else:
                    print(f"    [+] Logged metrics for Port {port} Scenario: {rule['name']} ({dur}s)")
            except Exception as e:
                thumb_cell.value = f"(Image failed: {e})"

            ws_main.row_dimensions[current_main_row].height = ROW_H
            current_main_row += 1
            rows_written_in_current_file += 1

# --- 8. EXPORT ENGINE ---
file_path = output_dir / REPORT_PART_FILE_TEMPLATE.format(batch_idx=batch_index)

if batch_index == 1 and rows_written_in_current_file < MAX_ROWS_PER_FILE:
    file_path = output_dir / f"{base_filename}_{DATE_SUFFIX}.xlsx"

wb.save(file_path)
generated_files.append(file_path)

# --- 9. CONDITIONAL AUTO-ZIP COMPRESSION MATRIX ENGINE ---
final_output_path = file_path # Default fallback for auto-open logic

if len(generated_files) > 1:
    zip_output_path = downloads_path / FINAL_ZIP_ARCHIVE_NAME
    print(f"\n[+] Multiple files generated ({len(generated_files)} parts). Packing into a secure ZIP container...")
    
    with zipfile.ZipFile(zip_output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in generated_files:
            zipf.write(file, arcname=file.name)
            os.remove(file)
            
    print(f"[+] Success! Consolidated archives compiled into:\n    {zip_output_path}")
    final_output_path = zip_output_path
    try:
        os.rmdir(output_dir)
    except Exception:
        pass
else:
    if generated_files:
        print(f"\n[+] Success! All entries processed. Output saved cleanly to:\n    {file_path}")

# --- 10. PERFORMANCE ANALYTICS & AUTO-OPEN FINALE ---
elapsed_seconds = time.time() - script_start_time
minutes = int(elapsed_seconds // 60)
seconds = int(elapsed_seconds % 60)

print(f"\n=================================================================")
print(f" BATCH PROCESSING COMPLETED SUCCESSFULLY")
print(f"=================================================================")
print(f" -> Total Run Duration: {minutes}m {seconds}s ({round(elapsed_seconds, 2)} total seconds)")
print(f" -> Files Output Group: {len(generated_files)} Workbook Sheet Chunk(s)")
print(f"=================================================================\n")

# AUTO-OPEN THE RESULTS DIRECTORY
try:
    print("[*] Opening results folder...")
    folder_to_open = final_output_path.parent
    if platform.system() == "Windows":
        os.startfile(folder_to_open)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", folder_to_open])
    else:
        subprocess.Popen(["xdg-open", folder_to_open])
except Exception as e:
    pass