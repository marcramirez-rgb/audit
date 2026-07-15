"""Camera analytics engine: fetching, parsing, rendering, and report generation.

No input()/print()/exit() here -- this module is a plain library used by both
the CLI (combined.py) and the GUI (gui_app.py). Callers pass in credentials and
a log/progress callback instead of relying on module-level globals or stdout.
"""

import concurrent.futures
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
from pathlib import Path

import requests
import urllib3
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from PIL import Image as PILImage, ImageDraw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

STRICT_TIMEOUT = (3.05, 5.0)
MAX_WORKER_THREADS = 15

THUMB_H = 250
ROW_H = 190

CAMERA_CONFIGS = [
    {"port": "5010", "position": "CENTER", "color": "E5F5F5"},  # LVT Light
    {"port": "5015", "position": "LEFT",   "color": "00A19A"},  # LVT Normal
    {"port": "5020", "position": "RIGHT",  "color": "00726E"}   # LVT Dark
]


def default_output_dir():
    downloads_path = Path.home() / "Downloads"
    output_dir = downloads_path / "Camera_Reports_Master"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def classify_manufacturer(raw_value):
    """Returns 'AXIS', 'HIKVISION', 'MIXED', or None if the value doesn't clearly
    match any of those. Never silently defaults -- an unrecognized value (blank,
    misspelled, wrong CSV column mapped in) must be caught by the caller instead
    of routed to the wrong API. "LVT" is a rebranded/OEM camera line running
    Hikvision firmware -- same ISAPI endpoints, same credentials -- so it's
    treated as HIKVISION here. "MIXED" means the physical unit's three camera
    positions aren't all the same vendor (e.g. two Hikvision, one Axis) -- see
    process_camera_row's MIXED branch for how that's actually resolved per port."""
    value = (raw_value or "").strip().lower()
    if "mixed" in value or "both" in value:
        return "MIXED"
    if "axis" in value:
        return "AXIS"
    if "hik" in value or "lvt" in value:
        return "HIKVISION"
    return None


def dedupe_camera_rows(rows):
    """Collapse rows sharing the same IP into a single device row.

    If an IP appears multiple times with more than one recognized vendor, the
    resulting row is marked as MANUFACTURER='mixed' so the engine can probe each
    port against both Axis and Hikvision APIs.
    """
    grouped = {}
    order = []
    passthrough = []
    for row in rows:
        ip = (row.get("IP", "") or "").strip()
        if not ip:
            passthrough.append(row)
            continue
        if ip not in grouped:
            grouped[ip] = []
            order.append(ip)
        grouped[ip].append(row)

    deduped = []
    for ip in order:
        group = grouped[ip]
        if len(group) == 1:
            deduped.append(group[0])
            continue

        classes = {classify_manufacturer(r.get("MANUFACTURER", "")) for r in group}
        classes.discard(None)
        base = dict(group[0])
        if len(classes) > 1:
            base["MANUFACTURER"] = "mixed"
        deduped.append(base)

    return deduped + passthrough


# --- OBJECT-ORIENTED ARCHITECTURE (OOP) ---

class CameraHandler:
    """Base class for camera manufacturers."""
    def __init__(self, ip, user, password):
        self.ip = ip
        self.auth_strategies = [HTTPDigestAuth(user, password), HTTPBasicAuth(user, password)]

    def fetch_snapshot(self, session, port):
        raise NotImplementedError

    def fetch_analytics(self, session, port):
        raise NotImplementedError

    def parse_analytics(self, data, img_w, img_h):
        raise NotImplementedError


class HikvisionHandler(CameraHandler):
    def __init__(self, ip, user, password):
        super().__init__(ip, user, password)
        self.fallback_dim = (1920, 1080)

    def fetch_snapshot(self, session, port):
        img_url = f"http://{self.ip}:{port}/ISAPI/Streaming/channels/101/picture"
        last_err = "All authentication attempts failed"
        saw_401 = False
        saw_other = False
        for auth in self.auth_strategies:
            try:
                with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, stream=True, verify=False) as response:
                    if response.status_code == 200:
                        return PILImage.open(BytesIO(response.content)), img_url, None, False
                    if response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                    last_err = f"HTTP {response.status_code}: {response.text[:150]}"
            except requests.exceptions.RequestException as e:
                saw_other = True
                last_err = f"{type(e).__name__}: {e}"
                continue
        # auth_rejected is only True when every attempt cleanly got HTTP 401 -- a
        # timeout/connection error/other status means we can't be confident it's a
        # credential problem, so remaining ports on this device still get tried.
        auth_rejected = saw_401 and not saw_other
        return None, img_url, last_err, auth_rejected

    # Known Hikvision thermal/bi-spectrum model prefix. Deliberately NOT matching on a
    # bare "-T" -- Hikvision's regular optical turret cameras use "T" in their model
    # codes too (e.g. DS-2CD2T...), which would false-positive on a normal fleet.
    THERMAL_MODEL_MARKERS = ["2TD"]

    def _detect_device_model(self, session, port):
        """Best-effort only: not verified against a real failing device (unlike the
        Axis Perimeter Defender detection, which was confirmed via packet capture).
        Uses Hikvision's standard deviceInfo endpoint to flag likely thermal/specialty
        hardware when the normal behaviorRule analytics endpoint doesn't work."""
        info_url = f"http://{self.ip}:{port}/ISAPI/System/deviceInfo"
        for auth in self.auth_strategies:
            try:
                response = session.get(info_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
                if response.status_code == 200:
                    try:
                        root = ET.fromstring(response.text)
                    except ET.ParseError:
                        return None
                    for elem in root.iter():
                        if elem.tag.split('}')[-1] == 'model' and elem.text:
                            return elem.text.strip()
                    return None
            except requests.exceptions.RequestException:
                continue
        return None

    def fetch_analytics(self, session, port):
        last_status = None
        last_snippet = None
        saw_401 = False
        saw_other = False
        for channel_id in [1, 2]:
            rule_url = f"http://{self.ip}:{port}/ISAPI/Intelligent/channels/{channel_id}/behaviorRule/1"
            for auth in self.auth_strategies:
                try:
                    response = session.get(rule_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
                    last_status = response.status_code
                    if response.status_code == 200:
                        temp_xml = response.text
                        if temp_xml and ("positionX" in temp_xml or "RegionCoordinates" in temp_xml):
                            return temp_xml, rule_url, None, False
                        last_snippet = (temp_xml[:200] + "...") if temp_xml and len(temp_xml) > 200 else (temp_xml or "(empty body)")
                        saw_other = True
                    elif response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                except requests.exceptions.RequestException as e:
                    saw_other = True
                    last_snippet = f"{type(e).__name__}: {e}"
                    continue
        auth_rejected = saw_401 and not saw_other
        if last_status is None:
            diag = f"No response from any channel/auth combo (last error: {last_snippet})"
        else:
            diag = f"Last HTTP {last_status} response did not contain expected rule tags. Body snippet: {last_snippet}"

        model = self._detect_device_model(session, port)
        if model and any(marker in model.upper() for marker in self.THERMAL_MODEL_MARKERS):
            return None, f"http://{self.ip}:{port}/ISAPI/Intelligent/channels/[1,2]/behaviorRule/1", \
                (f"SPECIAL CASE: Device model '{model}' looks like a thermal/specialty unit -- standard behaviorRule "
                 f"analytics API may not apply. Requires manual review, not confirmed against real device traffic. | {diag}"), auth_rejected
        return None, f"http://{self.ip}:{port}/ISAPI/Intelligent/channels/[1,2]/behaviorRule/1", f"No active analytic perimeters found | {diag}", auth_rejected

    def parse_analytics(self, xml_data, img_w, img_h):
        namespaces = {'ns': 'http://www.std-cgi.com/ver20/XMLSchema'}
        try:
            root = ET.fromstring(xml_data)
            rules = root.findall('.//ns:RuleInfo', namespaces)
        except Exception:
            rules = []

        if not rules:
            return [{"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]

        parsed_rules = []
        for rule in rules:
            rule_name = rule.find('ns:ruleName', namespaces).text

            event_type_raw = rule.find('ns:eventType', namespaces)
            event_type = event_type_raw.text if event_type_raw is not None else "Unknown"
            if "field" in event_type.lower() or "intrusion" in event_type.lower():
                event_type = "Intrusion Detection"
            elif "line" in event_type.lower() or "cross" in event_type.lower():
                event_type = "Line Crossing"

            duration_val = "0"
            possible_tags = ['.//ns:FieldDetectionParam/ns:durationTime', './/ns:FieldDetectionParam/ns:timeDuration', './/ns:LineCrossing/ns:duration', './/ns:Intrusion/ns:duration', './/ns:durationTime', './/ns:timeDuration', './/ns:duration', './/ns:alarmDelay']
            found_values = [rule.find(tag, namespaces).text.strip() for tag in possible_tags if rule.find(tag, namespaces) is not None and rule.find(tag, namespaces).text]
            if found_values:
                valid_seconds = [v for v in found_values if v != "100" and v != "0"]
                duration_val = valid_seconds[0] if valid_seconds else found_values[0]

            target_node = rule.find('.//ns:FieldDetectionParam/ns:detectionTarget', namespaces)
            if target_node is None:
                target_node = rule.find('.//ns:detectionTarget', namespaces) or rule.find('.//ns:TargetType', namespaces)
            target_detection = target_node.text.capitalize() if (target_node is not None and target_node.text) else "All Targets"

            region_lists = rule.findall('.//ns:RegionCoordinatesList', namespaces)
            polygons = []
            for region in region_lists:
                vertices = []
                for coord in region.findall('ns:RegionCoordinates', namespaces):
                    raw_x = float(coord.find('ns:positionX', namespaces).text)
                    raw_y = float(coord.find('ns:positionY', namespaces).text)

                    # Hikvision positionX/positionY use a 0..1000 space with (0,0)
                    # at the top-left. Map directly and clamp to bounds. Previous
                    # behavior inverted axes which caused mirrored overlays.
                    pixel_x = int((raw_x / 1000.0) * img_w)
                    pixel_y = int((raw_y / 1000.0) * img_h)
                    pixel_y = img_h - 1 - pixel_y
                    pixel_x = max(0, min(img_w - 1, pixel_x))
                    pixel_y = max(0, min(img_h - 1, pixel_y))
                    vertices.append((pixel_x, pixel_y))

                if vertices:
                    polygons.append(vertices)

            if not polygons:
                continue

            for zone_index, polygon in enumerate(polygons, start=1):
                parsed_rules.append({
                    "is_placeholder": False,
                    "name": rule_name if len(polygons) == 1 else f"{rule_name} [Zone {zone_index}]",
                    "type": event_type,
                    "duration": duration_val,
                    "target": target_detection,
                    "vertices": polygon
                })

        def normalize_vertices_key(verts):
            if not verts:
                return ()
            if isinstance(verts[0], (list, tuple)) and verts[0] and isinstance(verts[0][0], (list, tuple)):
                return tuple(tuple((int(x), int(y)) for x, y in polygon) for polygon in verts)
            return tuple((int(x), int(y)) for x, y in verts)

        seen = set()
        deduped = []
        for r in parsed_rules:
            verts_key = normalize_vertices_key(r.get("vertices", []))
            key = (r.get("name"), r.get("type"), r.get("target"), r.get("duration"), verts_key)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)

        if deduped:
            return deduped
        return [{"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]
        seen = set()
        deduped = []
        for r in parsed_rules:
            verts_key = tuple((int(x), int(y)) for x, y in r.get("vertices", []))
            key = (r.get("name"), r.get("type"), r.get("target"), r.get("duration"), verts_key)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)

        if deduped:
            return deduped
        return [{"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]


class AxisHandler(CameraHandler):
    def __init__(self, ip, user, password):
        super().__init__(ip, user, password)
        self.fallback_dim = (1280, 720)

    def fetch_snapshot(self, session, port):
        img_url = f"http://{self.ip}:{port}/axis-cgi/jpg/image.cgi?resolution=1280x720"
        last_err = "All authentication attempts failed"
        saw_401 = False
        saw_other = False
        for auth in self.auth_strategies:
            try:
                with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, stream=True, verify=False) as response:
                    if response.status_code == 200:
                        return PILImage.open(BytesIO(response.content)), img_url, None, False
                    if response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                    last_err = f"HTTP {response.status_code}: {response.text[:150]}"
            except requests.exceptions.RequestException as e:
                saw_other = True
                last_err = f"{type(e).__name__}: {e}"
                continue
        auth_rejected = saw_401 and not saw_other
        return None, img_url, last_err, auth_rejected

    def _detect_active_analytics_app(self, session, port):
        """Called only when the generic Object Analytics endpoint 404s. Checks the
        camera's installed-application list to identify what analytics engine (if
        any) is actually running -- e.g. AXIS Perimeter Defender on fixed thermal
        units, which is a separate ACAP app with no relation to Object Analytics."""
        list_url = f"http://{self.ip}:{port}/axis-cgi/applications/list.cgi"
        for auth in self.auth_strategies:
            try:
                response = session.get(list_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
                if response.status_code == 200:
                    try:
                        root = ET.fromstring(response.text)
                    except ET.ParseError:
                        return None
                    return [app.get("NiceName", app.get("Name", "Unknown App"))
                            for app in root.findall(".//application")
                            if app.get("Status") == "Running"]
            except requests.exceptions.RequestException:
                continue
        return None

    def fetch_analytics(self, session, port):
        control_url = f"http://{self.ip}:{port}/local/objectanalytics/control.cgi"
        payload = {"apiVersion": "1.2", "method": "getConfiguration"}
        last_err = "All authentication attempts failed"
        saw_404 = False
        saw_401 = False
        saw_other = False

        for auth in self.auth_strategies:
            try:
                response = session.post(control_url, auth=auth, json=payload, timeout=STRICT_TIMEOUT, verify=False)
                if response.status_code == 200:
                    json_data = response.json()
                    # Axis's local CGI APIs return HTTP 200 with an "error" body on failure
                    # (e.g. Object Analytics app not running/licensed, unsupported apiVersion)
                    # instead of a non-200 status. Without this check that error was silently
                    # treated as "zero scenarios configured".
                    if isinstance(json_data, dict) and "error" in json_data:
                        err = json_data["error"]
                        if isinstance(err, dict):
                            last_err = f"API error {err.get('code', '?')}: {err.get('message', json_data)}"
                        else:
                            last_err = f"API error: {err}"
                        saw_other = True
                        continue
                    return json_data, control_url, None, False
                else:
                    if response.status_code == 404:
                        saw_404 = True
                    elif response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                    last_err = f"HTTP Error Status {response.status_code}: {response.text[:200]}"
            except Exception as e:
                saw_other = True
                last_err = f"{type(e).__name__}: {e}"
                continue

        auth_rejected = saw_401 and not saw_other and not saw_404
        if saw_404:
            running_apps = self._detect_active_analytics_app(session, port)
            if running_apps:
                last_err = (f"SPECIAL CASE: This device has no AXIS Object Analytics app -- "
                            f"active analytics app instead: {', '.join(running_apps)}. "
                            f"Requires a separate integration, not a code failure.")
        return None, control_url, last_err, auth_rejected

    def parse_analytics(self, json_data, img_w, img_h):
        scenarios = json_data.get("data", {}).get("scenarios", [])
        if not scenarios:
            return [{"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]

        parsed_rules = []
        for scenario in scenarios:
            if scenario.get("is_placeholder"):
                parsed_rules.append({"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []})
                continue

            rule_name = scenario.get("name", f"Scenario {scenario.get('id')}")
            loiter_time = ""
            triggers = scenario.get("triggers", [])

            for trigger_obj in triggers:
                for condition in trigger_obj.get("conditions", []):
                    if condition.get("type") == "individualTimeInArea":
                        for d in condition.get("data", []):
                            if d.get("time"):
                                loiter_time = str(d["time"])
                                break

            if not loiter_time:
                for f_obj in scenario.get("filters", []):
                    if f_obj.get("type") == "timeShortLivedLimit":
                        loiter_time = str(f_obj.get("time", ""))
                        break

            duration_val = loiter_time if loiter_time else "0"
            classes = [oc.get("type", "") for oc in scenario.get("objectClassifications", []) if oc.get("type")]
            target_detection = ", ".join(c.capitalize() + " Detection" for c in classes) if classes else "Any Detection"

            rule_type_base = scenario.get("type", "")
            if triggers and not rule_type_base: rule_type_base = triggers[0].get("type", "Unknown")
            rule_type = "Intrusion Detection" if "area" in rule_type_base.lower() else rule_type_base.capitalize()

            vertices = []
            if triggers:
                raw_vertices = triggers[0].get("vertices", [])
                for pt in raw_vertices:
                    raw_x = float(pt[0])
                    raw_y = float(pt[1])
                    pixel_x = int(((raw_x + 1.0) / 2.0) * img_w)
                    pixel_y = int(((1.0 - raw_y) / 2.0) * img_h)
                    vertices.append((pixel_x, pixel_y))

            parsed_rules.append({
                "is_placeholder": False,
                "name": rule_name,
                "type": rule_type,
                "duration": duration_val,
                "target": target_detection,
                "vertices": vertices
            })

        return parsed_rules


# --- UNIVERSAL RENDERING ENGINE (JPEG COMPRESSION) ---

def render_overlay_image(camera_image, vertices, index, img_w, img_h, rule_type=""):
    snapshot_missing = (camera_image is None)
    if camera_image is not None:
        base_img = camera_image.copy().convert("RGBA")
    else:
        base_img = PILImage.new("RGBA", (img_w, img_h), (40, 40, 40, 255))

    brand_colors = [(0, 161, 154), (0, 114, 110), (229, 245, 245)]
    rgb_color = brand_colors[index % 3]

    overlay = PILImage.new("RGBA", base_img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    def _get_polygons(verts):
        if not verts:
            return []
        if isinstance(verts[0], (list, tuple)) and verts[0] and isinstance(verts[0][0], (list, tuple)):
            return verts
        return [verts]

    for polygon in _get_polygons(vertices):
        is_line_rule = any(keyword in rule_type.lower() for keyword in ["line", "cross", "fence"])
        polygon_to_draw = polygon
        try:
            if not is_line_rule and len(polygon) > 2:
                import math
                cx = sum([p[0] for p in polygon]) / len(polygon)
                cy = sum([p[1] for p in polygon]) / len(polygon)
                polygon_to_draw = sorted(polygon, key=lambda p: math.atan2(p[1] - cy, p[0] - cx))
        except Exception:
            polygon_to_draw = polygon

        if is_line_rule or len(polygon_to_draw) <= 2:
            draw.line(polygon_to_draw, fill=rgb_color + (255,), width=4)
        else:
            draw.polygon(polygon_to_draw, fill=rgb_color + (76,))
            draw.polygon(polygon_to_draw, outline=rgb_color + (255,), width=3)
        for (x, y) in polygon_to_draw:
            draw.ellipse((x - 6, y - 6, x + 6, y + 6), fill=(255, 255, 0, 255), outline=(0, 0, 0, 255), width=1)

    if snapshot_missing:
        try:
            from PIL import ImageFont
            font = ImageFont.load_default()
            text = "SNAPSHOT UNAVAILABLE"
            try:
                text_width, text_height = draw.textsize(text, font=font)
            except Exception:
                text_width, text_height = font.getsize(text)
            padding = 12
            text_box = [
                (base_img.width - text_width) // 2 - padding,
                (base_img.height - text_height) // 2 - padding,
                (base_img.width + text_width) // 2 + padding,
                (base_img.height + text_height) // 2 + padding,
            ]
            draw.rectangle(text_box, fill=(0, 0, 0, 200), outline=(255, 255, 255, 255), width=2)
            draw.text(((base_img.width - text_width) // 2, (base_img.height - text_height) // 2), text, fill=(255, 255, 255, 255), font=font)
        except Exception:
            draw.text((10, 10), "SNAPSHOT UNAVAILABLE", fill=(255, 255, 255, 255))

    final_img = PILImage.alpha_composite(base_img, overlay).convert("RGB")

    ratio = THUMB_H / final_img.height
    img_resized = final_img.resize((int(final_img.width * ratio), THUMB_H), PILImage.Resampling.LANCZOS)

    img_buf = BytesIO()
    img_resized.save(img_buf, format="JPEG", quality=85)
    img_buf.seek(0)

    return img_buf


def create_master_workbook():
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Camera Analytics"
    ws_main.views.sheetView[0].showGridLines = True

    ws_missed = wb.create_sheet(title="Missed Cameras")
    ws_missed.views.sheetView[0].showGridLines = True

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="1A1D27")
    failed_fill   = PatternFill("solid", start_color="A61C1C")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_main.merge_cells("A1:I1")
    ws_main["A1"] = f"Intelligent Analytics Master Report"
    ws_main["A1"].font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws_main["A1"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A1"].alignment = center_align
    ws_main.row_dimensions[1].height = 36

    ws_main.merge_cells("A2:I2")
    ws_main["A2"] = f"Batch Processing Timeline Context: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_main["A2"].font      = Font(name="Arial", size=10, color="888888")
    ws_main["A2"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A2"].alignment = center_align
    ws_main.row_dimensions[2].height = 20

    main_headers = ["Client Name", "Location", "Live Unit Serial", "Camera Position", "Rule Name", "Rule Type", "Target Detection", "Duration (s)", "Rule Visual Overlay Thumbnail"]
    for col, h in enumerate(main_headers, 1):
        cell = ws_main.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        cell.border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"), top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
    ws_main.row_dimensions[3].height = 22

    main_widths = [22, 24, 22, 16, 20, 20, 20, 14, 110]
    for i, w in enumerate(main_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = w

    ws_missed.merge_cells("A1:G1")
    ws_missed["A1"] = "Analytics Fetch Exception Audit Log"
    ws_missed["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_missed["A1"].fill      = PatternFill("solid", start_color="330000")
    ws_missed["A1"].alignment = center_align
    ws_missed.row_dimensions[1].height = 36

    missed_headers = ["Timestamp", "Client Name", "Location", "Live Unit Serial", "Camera Port", "Target Endpoint", "Failure Reason"]
    for col, h in enumerate(missed_headers, 1):
        cell = ws_missed.cell(row=2, column=col, value=h)
        cell.font = header_font; cell.fill = failed_fill; cell.alignment = center_align
    ws_missed.row_dimensions[2].height = 22
    missed_widths = [22, 22, 24, 22, 15, 65, 45]
    for i, w in enumerate(missed_widths, 1):
        ws_missed.column_dimensions[get_column_letter(i)].width = w

    ws_main.freeze_panes = "A4"
    ws_missed.freeze_panes = "A3"

    return wb, ws_main, ws_missed


# --- MULTI-THREADED WORKER FUNCTION ---

def process_camera_row(args):
    row_idx, row_data, sess, credentials = args
    client_name = row_data.get("CLIENT_NM", "").strip()
    location = row_data.get("LOCATION_NM", "").strip()
    serial = row_data.get("LIVE_UNIT_SERIAL_NM", "").strip()
    ip = row_data.get("IP", "").strip()

    user_axis, pass_axis = credentials.get("AXIS_USER"), credentials.get("AXIS_PASS")
    user_hik, pass_hik = credentials.get("HIK_USER"), credentials.get("HIK_PASS")

    raw_mfg = row_data.get("MANUFACTURER", "")
    mfg_class = classify_manufacturer(raw_mfg)
    api_target_str = mfg_class or "UNKNOWN"

    results = {
        "ip": ip, "idx": row_idx, "client": client_name, "loc": location,
        "target": api_target_str, "logs": [], "main": [], "missed": []
    }

    if not ip:
        return results

    if mfg_class is None:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        results["missed"].append([timestamp, client_name, location, serial, "N/A", "N/A",
                                   f"Unrecognized MANUFACTURER value: '{raw_mfg}' (must contain 'axis' or 'hik') -- "
                                   f"analytics skipped, snapshot attempted with both known API formats"])

        # We don't know the vendor, so we don't know which snapshot URL format applies.
        # Try both rather than giving up on getting at least a photo for this row.
        # Guard on credentials being non-empty -- constructing a handler with a None
        # user/password doesn't fail cleanly, it raises a TypeError mid-request when
        # HTTPBasicAuth/HTTPDigestAuth try to encode it.
        probe_handlers = []
        if user_axis and pass_axis:
            probe_handlers.append(AxisHandler(ip, user_axis, pass_axis))
        if user_hik and pass_hik:
            probe_handlers.append(HikvisionHandler(ip, user_hik, pass_hik))

        for cam in CAMERA_CONFIGS:
            port = cam["port"]
            pos = cam["position"]
            bg_color = cam["color"]
            results["logs"].append(f" -> Testing {pos} Interface Port: {port} (unrecognized manufacturer, snapshot-only)...")

            camera_image, img_w, img_h = None, 1280, 720
            for probe in probe_handlers:
                camera_image, snap_url, snap_err, _ = probe.fetch_snapshot(sess, port)
                if camera_image is not None:
                    img_w, img_h = camera_image.size
                    break
            if camera_image is None:
                results["logs"].append(f"    [!] Warning: Could not fetch snapshot with any known API format. Defaulting to {img_w}x{img_h} canvas.")

            img_buf = render_overlay_image(camera_image, [], 0, img_w, img_h, "")
            results["main"].append({
                "data": [client_name, location, serial, pos, "Unrecognized Manufacturer", "N/A", "Analytics Skipped", "N/A"],
                "bg": bg_color, "img": img_buf, "err": ""
            })
            if camera_image is not None:
                camera_image.close()

        return results

    if mfg_class == "MIXED":
        # A single IP whose three physical camera positions aren't all the same
        # vendor (e.g. two Hikvision, one Axis). Unlike the unrecognized-manufacturer
        # branch above, we want real per-port analytics here, not just a snapshot --
        # so each port gets a cheap snapshot probe per vendor first (one request per
        # auth strategy) to see which vendor actually answers, then only that vendor's
        # full analytics fetch runs for that port. NOTE: this branch expects the caller
        # to have already collected BOTH AXIS_USER/PASS and HIK_USER/PASS -- run_audit.py
        # does this for a "mixed" row, but combined.py/gui_app.py compute their own
        # needs_axis/needs_hik pre-flight check independently and don't know about
        # "MIXED" yet, so a MIXED row run through those front ends may not get prompted
        # for both credential sets.
        axis_handler = AxisHandler(ip, user_axis, pass_axis) if (user_axis and pass_axis) else None
        hik_handler = HikvisionHandler(ip, user_hik, pass_hik) if (user_hik and pass_hik) else None

        # Once a vendor's credentials get a clean 401 on this device, stop trying
        # them on the device's other ports -- same account-lockout-avoidance
        # reasoning as the single-vendor path below, just tracked per vendor
        # since a MIXED device involves two independent logins, not one.
        axis_dead = axis_handler is None
        hik_dead = hik_handler is None
        center_axis_only = None

        for cam in CAMERA_CONFIGS:
            port = cam["port"]
            pos = cam["position"]
            bg_color = cam["color"]
            results["logs"].append(f" -> Testing {pos} Interface Port: {port} (mixed-vendor unit, probing)...")

            # If the center camera proves to be Axis-only, skip Axis probes on side ports.
            if pos == "CENTER" and axis_handler and hik_handler and center_axis_only is None:
                axis_probe_ok = False
                hik_probe_ok = False

                probe_img, _, _, probe_auth_rejected = axis_handler.fetch_snapshot(sess, port)
                if probe_img is not None:
                    probe_img.close()
                    axis_probe_ok = True
                elif probe_auth_rejected:
                    axis_dead = True

                probe_img, _, _, probe_auth_rejected = hik_handler.fetch_snapshot(sess, port)
                if probe_img is not None:
                    probe_img.close()
                    hik_probe_ok = True
                elif probe_auth_rejected:
                    hik_dead = True

                if axis_probe_ok and not hik_probe_ok:
                    center_axis_only = True
                    results["logs"].append("    [>] Mixed unit center port is Axis-only; skipping Axis probes on non-center ports.")
                elif hik_probe_ok and not axis_probe_ok:
                    center_axis_only = False
                    results["logs"].append("    [>] Mixed unit center port is Hikvision-only; non-center ports will still be probed normally.")

            handler, vendor_label, camera_image = None, None, None
            probe_errors = []

            try_axis = not axis_dead and not (center_axis_only is True and pos != "CENTER")
            if try_axis:
                img, snap_url, snap_err, auth_rejected = axis_handler.fetch_snapshot(sess, port)
                if img is not None:
                    handler, vendor_label, camera_image = axis_handler, "Axis", img
                else:
                    if auth_rejected:
                        axis_dead = True
                        results["logs"].append(f"    [!] Axis credentials flatly rejected on Port {port} -- won't retry Axis on this device's remaining ports.")
                    probe_errors.append(f"Axis: {snap_err}")
            elif center_axis_only is True:
                results["logs"].append(f"    [>] Skipping Axis probe on Port {port} because center port is Axis-only.")

            if handler is None and not hik_dead:
                img, snap_url, snap_err, auth_rejected = hik_handler.fetch_snapshot(sess, port)
                if img is not None:
                    handler, vendor_label, camera_image = hik_handler, "Hikvision", img
                else:
                    if auth_rejected:
                        hik_dead = True
                        results["logs"].append(f"    [!] Hikvision credentials flatly rejected on Port {port} -- won't retry Hikvision on this device's remaining ports.")
                    probe_errors.append(f"Hikvision: {snap_err}")

            if handler is None:
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                reason = " | ".join(probe_errors) if probe_errors else "no credentials provided for either vendor"
                results["missed"].append([timestamp, client_name, location, serial, f"{pos} ({port}) [MIXED]", "N/A",
                                           f"Neither vendor produced a snapshot on this port -- {reason}"])
                img_buf = render_overlay_image(None, [], 0, 1280, 720, "")
                results["main"].append({
                    "data": [client_name, location, serial, pos, "Unresolved Vendor", "N/A", "Neither Vendor Responded", "N/A"],
                    "bg": bg_color, "img": img_buf, "err": ""
                })
                continue

            img_w, img_h = camera_image.size
            payload_data, req_url, err_msg, analytics_auth_rejected = handler.fetch_analytics(sess, port)
            if analytics_auth_rejected:
                if vendor_label == "Axis":
                    axis_dead = True
                else:
                    hik_dead = True

            pos_label = f"{pos} [{vendor_label}]"

            if payload_data is None:
                err_msg = err_msg or "Unauthorized Connection (All auth variations failed)"
                is_special_case = err_msg.startswith("SPECIAL CASE:")
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                tag = "[SPECIAL CASE]" if is_special_case else ""
                results["missed"].append([timestamp, client_name, location, serial, f"{pos_label} ({port}) {tag}".strip(), req_url, err_msg])
                placeholder_name = "Non-Standard Analytics App (See Missed Tab)" if is_special_case else "Analytics Fetch Failed (Check Missed Tab)"
                rules = [{"is_placeholder": True, "name": placeholder_name, "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]
            else:
                rules = handler.parse_analytics(payload_data, img_w, img_h)

            for index, rule in enumerate(rules):
                try:
                    img_buf = render_overlay_image(camera_image, rule["vertices"], index, img_w, img_h, rule["type"])
                    display_name = rule["name"]
                    if camera_image is None:
                        display_name = f"Snapshot Failed: {display_name}"
                    results["main"].append({
                        "data": [client_name, location, serial, pos_label, display_name, rule["type"], rule["target"], rule["duration"]],
                        "bg": bg_color, "img": img_buf, "err": ""
                    })
                    if rule.get("is_placeholder"):
                        results["logs"].append(f"    [+] Logged snapshot row for Port {port} (No rules active or fetch failed) -- resolved as {vendor_label}")
                    else:
                        results["logs"].append(f"    [+] Logged metrics for Port {port} Scenario: {rule['name']} ({rule['duration']}s) -- resolved as {vendor_label}")
                except Exception as e:
                    results["main"].append({
                        "data": [client_name, location, serial, pos_label, rule["name"], rule["type"], rule["target"], rule["duration"]],
                        "bg": bg_color, "img": None, "err": f"(Image failed: {e})"
                    })

            if camera_image is not None:
                camera_image.close()

        return results

    is_axis = (mfg_class == "AXIS")
    handler = AxisHandler(ip, user_axis, pass_axis) if is_axis else HikvisionHandler(ip, user_hik, pass_hik)

    for cam in CAMERA_CONFIGS:
        port = cam["port"]
        pos = cam["position"]
        bg_color = cam["color"]
        results["logs"].append(f" -> Testing {pos} Interface Port: {port}...")

        camera_image, snap_url, snap_err, snap_auth_rejected = handler.fetch_snapshot(sess, port)
        if camera_image is None:
            img_w, img_h = handler.fallback_dim
            results["logs"].append(f"    [!] Warning: Failed to fetch stream snapshot ({snap_err}). Defaulting to {img_w}x{img_h} canvas.")
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            results["missed"].append([timestamp, client_name, location, serial, f"{pos} ({port}) [SNAPSHOT]", snap_url, snap_err])
        else:
            img_w, img_h = camera_image.size

        payload_data, req_url, err_msg, analytics_auth_rejected = handler.fetch_analytics(sess, port)
        if payload_data is not None and camera_image is None:
            results["logs"].append("    [!] Analytics retrieved but snapshot unavailable; rendering overlay on placeholder canvas.")

        if payload_data is None:
            err_msg = err_msg or "Unauthorized Connection (All auth variations failed)"
            is_special_case = err_msg.startswith("SPECIAL CASE:")
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            tag = "[SPECIAL CASE]" if is_special_case else ""
            results["missed"].append([timestamp, client_name, location, serial, f"{pos} ({port}) {tag}".strip(), req_url, err_msg])
            placeholder_name = "Non-Standard Analytics App (See Missed Tab)" if is_special_case else "Analytics Fetch Failed (Check Missed Tab)"
            rules = [{"is_placeholder": True, "name": placeholder_name, "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]
        else:
            rules = handler.parse_analytics(payload_data, img_w, img_h)

        for index, rule in enumerate(rules):
            try:
                img_buf = render_overlay_image(camera_image, rule["vertices"], index, img_w, img_h, rule["type"])
                results["main"].append({
                    "data": [client_name, location, serial, pos, rule["name"], rule["type"], rule["target"], rule["duration"]],
                    "bg": bg_color, "img": img_buf, "err": ""
                })
                if rule.get("is_placeholder"):
                    results["logs"].append(f"    [+] Logged snapshot row for Port {port} (No rules active or fetch failed)")
                else:
                    results["logs"].append(f"    [+] Logged metrics for Port {port} Scenario: {rule['name']} ({rule['duration']}s)")
            except Exception as e:
                results["main"].append({
                    "data": [client_name, location, serial, pos, rule["name"], rule["type"], rule["target"], rule["duration"]],
                    "bg": bg_color, "img": None, "err": f"(Image failed: {e})"
                })

        if camera_image is not None:
            camera_image.close()

        # If credentials were flatly rejected (clean 401s, not timeouts/other errors) on
        # this port, the same login will fail identically on this device's other ports --
        # retrying them just multiplies failed-auth attempts and risks tripping the
        # device's account lockout. Skip the rest of this camera and note why.
        if snap_auth_rejected or analytics_auth_rejected:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            results["logs"].append(f"    [!] Authentication flatly rejected on Port {port} -- skipping remaining ports on this device to avoid triggering an account lockout.")
            results["missed"].append([timestamp, client_name, location, serial, "REMAINING PORTS SKIPPED", "N/A",
                                       f"Authentication rejected on port {port} -- remaining ports skipped to avoid repeated failed-login attempts against the same device."])
            break

    return results


def run_batch(camera_rows, credentials, output_dir, base_filename, log_cb=None, progress_cb=None):
    """Runs the full batch: spins up the thread pool, processes every row, writes
    the Excel report, and returns the saved Path.

    credentials: dict with any of AXIS_USER/AXIS_PASS/HIK_USER/HIK_PASS.
    log_cb(str): called for every line of progress output (defaults to a no-op).
    progress_cb(done, total): called after each row finishes (defaults to a no-op).
    """
    log = log_cb or (lambda _msg: None)
    report_progress = progress_cb or (lambda _done, _total: None)

    script_start_time = time.time()
    date_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    master_report_file = f"{base_filename}_Master_{date_suffix}.xlsx"

    current_main_row = 4
    current_missed_row = 3

    wb, ws_main, ws_missed = create_master_workbook()

    normal_font = Font(name="Arial", size=10)
    bold_font   = Font(name="Arial", bold=True, size=10)
    white_normal_font = Font(name="Arial", size=10, color="FFFFFF")
    white_bold_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    left_align  = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"), top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

    session = requests.Session()
    retries = Retry(total=2, backoff_factor=0.5, status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries, pool_connections=MAX_WORKER_THREADS, pool_maxsize=MAX_WORKER_THREADS)
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    total_cameras = len(camera_rows)
    log("[+] Spinning up multi-threaded batch engine...")

    # Never spin up more threads than there are rows to process (e.g. single-test mode
    # doesn't need a 15-thread pool for 1 camera).
    active_worker_count = max(1, min(MAX_WORKER_THREADS, total_cameras))

    done_count = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=active_worker_count) as executor:
        worker_args = [(i, row, session, credentials) for i, row in enumerate(camera_rows, 1)]

        for res in executor.map(process_camera_row, worker_args):
            if not res["ip"]:
                done_count += 1
                report_progress(done_count, total_cameras)
                continue

            log(f"\n#################################################################")
            log(f" INGESTING HOST NODE [{res['idx']}/{total_cameras}]: {res['client']} - {res['loc']} ({res['ip']}) [{res['target']}]")
            log(f"#################################################################")
            for line in res["logs"]:
                log(line)

            for missed_row in res["missed"]:
                ws_missed.append(missed_row)
                for col in range(1, 8):
                    ws_missed.cell(row=current_missed_row, column=col).border = thin_border
                    ws_missed.cell(row=current_missed_row, column=col).font = normal_font
                current_missed_row += 1

            for main_row in res["main"]:
                bg = main_row["bg"]
                row_fill = PatternFill("solid", start_color=bg)
                active_normal_font = white_normal_font if bg == "00726E" else normal_font
                active_bold_font = white_bold_font if bg == "00726E" else bold_font

                for col_idx, val in enumerate(main_row["data"], 1):
                    c = ws_main.cell(row=current_main_row, column=col_idx, value=val)
                    c.font = active_bold_font if col_idx == 4 else active_normal_font
                    c.fill = row_fill
                    c.alignment = center_align if col_idx in [3, 4, 7, 8] else left_align
                    c.border = thin_border

                thumb_cell = ws_main.cell(row=current_main_row, column=9, value=main_row["err"])
                thumb_cell.fill = row_fill; thumb_cell.border = thin_border

                if main_row["img"]:
                    xl_img = OpenpyxlImage(main_row["img"])
                    xl_img.anchor = TwoCellAnchor(editAs="oneCell", _from=AnchorMarker(col=8, colOff=0, row=current_main_row - 1, rowOff=0), to=AnchorMarker(col=9, colOff=0, row=current_main_row, rowOff=0))
                    ws_main.add_image(xl_img)

                ws_main.row_dimensions[current_main_row].height = ROW_H
                current_main_row += 1

            done_count += 1
            report_progress(done_count, total_cameras)

    final_output_path = Path(output_dir) / master_report_file
    wb.save(final_output_path)

    elapsed_seconds = time.time() - script_start_time
    minutes = int(elapsed_seconds // 60)
    seconds = int(elapsed_seconds % 60)

    log(f"\n=================================================================")
    log(f" MASTER BATCH PROCESSING COMPLETED SUCCESSFULLY")
    log(f"=================================================================")
    log(f" -> Total Run Duration: {minutes}m {seconds}s ({round(elapsed_seconds, 2)} total seconds)")
    log(f" -> Threads Utilized: {active_worker_count}")
    log(f" -> Master Report Saved: {final_output_path.name}")
    log(f"=================================================================\n")

    return final_output_path
