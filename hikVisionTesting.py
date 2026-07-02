import xml.etree.ElementTree as ET
import requests
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
import urllib3
import re
import os
import time
import zipfile
from io import BytesIO
from pathlib import Path
from datetime import datetime
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from PIL import Image as PILImage, ImageDraw

# --- GUI / Environment Imports ---
import tkinter as tk
from tkinter import filedialog
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Start the global execution benchmark timer
script_start_time = time.time()

# Suppress self-signed certificate warnings 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- 1. SET UP INFRASTRUCTURE PATHING & ENVIRONMENT CREDENTIALS ---
USER = os.getenv("HIK_USER", "admin") 
PASSWORD = os.getenv("HIK_PASSWORD")

# Secure Password Prompt Fallback
if not PASSWORD:
    print("[!] Security Notice: HIK_PASSWORD environment variable not found.")
    PASSWORD = input("Please enter the camera password to continue: ").strip()
    if not PASSWORD:
        print("[-] Error: Password cannot be blank. Exiting.")
        exit()

STRICT_TIMEOUT = (3.05, 5.0)

THUMB_H = 360  
ROW_H   = 275  
MAX_ROWS_PER_FILE = 100 

downloads_path = Path.home() / "Downloads"
output_dir = downloads_path / "Camera_Reports_Batch"
output_dir.mkdir(parents=True, exist_ok=True)

# Lens-specific port configurations (Using Kingfisher Brand Hex Codes)
CAMERA_CONFIGS = [
    {"port": "5010", "position": "CENTER", "color": "E5F5F5"},  # Kingfisher Light
    {"port": "5015", "position": "LEFT",   "color": "00A19A"},  # Kingfisher Normal
    {"port": "5020", "position": "RIGHT",  "color": "00726E"}   # Kingfisher Dark
]

# --- 2. MODULAR WORKER FUNCTIONS ---

def fetch_camera_snapshot(session, ip, port, auth_strategies):
    """Fetches the background stream snapshot using the active session."""
    img_url = f"http://{ip}:{port}/ISAPI/Streaming/channels/101/picture"
    for auth in auth_strategies:
        try:
            with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, stream=True) as response:
                if response.status_code == 200:
                    return PILImage.open(BytesIO(response.content))
        except requests.exceptions.RequestException:
            continue
    return None

def fetch_active_rules(session, ip, port, auth_strategies):
    """Probes channels 1 and 2 to find active analytic perimeters."""
    for channel_id in [1, 2]:
        rule_url = f"http://{ip}:{port}/ISAPI/Intelligent/channels/{channel_id}/behaviorRule/1"
        for auth in auth_strategies:
            try:
                response = session.get(rule_url, auth=auth, timeout=STRICT_TIMEOUT)
                if response.status_code == 200:
                    temp_xml = response.text
                    if temp_xml and ("positionX" in temp_xml or "RegionCoordinates" in temp_xml):
                        print(f"    [+] Successfully bound active rules on Channel {channel_id} (Port {port})")
                        return temp_xml, rule_url
            except requests.exceptions.RequestException:
                continue
    return None, ""

def parse_analytics_xml(xml_data, img_w, img_h):
    """Parses Hikvision XML payloads and extracts normalized coordinate matrices."""
    namespaces = {'ns': 'http://www.std-cgi.com/ver20/XMLSchema'}
    parsed_rules = []
    
    try:
        root = ET.fromstring(xml_data)
        rules = root.findall('.//ns:RuleInfo', namespaces)
    except Exception as e:
        print(f"    [!] XML Parsing Error: {e}")
        return []

    for rule in rules:
        rule_name = rule.find('ns:ruleName', namespaces).text
        
        event_type_raw = rule.find('ns:eventType', namespaces)
        event_type = event_type_raw.text if event_type_raw is not None else "Unknown"
        if "field" in event_type.lower() or "intrusion" in event_type.lower():
            event_type = "Intrusion Detection"
        elif "line" in event_type.lower() or "cross" in event_type.lower():
            event_type = "Line Crossing"

        # Duration parsing
        duration_val = "0"
        possible_tags = [
            './/ns:FieldDetectionParam/ns:durationTime', './/ns:FieldDetectionParam/ns:timeDuration',
            './/ns:LineCrossing/ns:duration', './/ns:Intrusion/ns:duration',
            './/ns:durationTime', './/ns:timeDuration', './/ns:duration', './/ns:alarmDelay'
        ]
        found_values = [rule.find(tag, namespaces).text.strip() for tag in possible_tags if rule.find(tag, namespaces) is not None and rule.find(tag, namespaces).text]
        if found_values:
            valid_seconds = [v for v in found_values if v != "100" and v != "0"]
            duration_val = valid_seconds[0] if valid_seconds else found_values[0]

        # Target classification
        target_node = rule.find('.//ns:FieldDetectionParam/ns:detectionTarget', namespaces)
        if target_node is None:
            target_node = rule.find('.//ns:detectionTarget', namespaces) or rule.find('.//ns:TargetType', namespaces)
        target_detection = target_node.text.capitalize() if (target_node is not None and target_node.text) else "All Targets"

        # Coordinate parsing with 180-degree web UI inversion correction
        region = rule.find('.//ns:RegionCoordinatesList', namespaces)
        if region is None:
            continue
            
        vertices = []
        for coord in region.findall('ns:RegionCoordinates', namespaces):
            raw_x = float(coord.find('ns:positionX', namespaces).text)
            raw_y = float(coord.find('ns:positionY', namespaces).text)
            
            pixel_x = int((raw_x / 1000.0) * img_w)
            pixel_y = int((raw_y / 1000.0) * img_h)
            
            # Correct the 180-degree web UI rotation inversion
            pixel_x = img_w - pixel_x
            pixel_y = img_h - pixel_y
            
            vertices.append((pixel_x, pixel_y))
            
        if not vertices:
            print(f"    [!] Warning: No vertices extracted for rule {rule_name}, skipping.")
            continue
            
        parsed_rules.append({
            "name": rule_name,
            "type": event_type,
            "duration": duration_val,
            "target": target_detection,
            "vertices": vertices
        })
        
    return parsed_rules

def render_overlay_image(camera_image, vertices, index, img_w, img_h):
    """Draws geometric overlays on the snapshot using Pillow."""
    if camera_image is not None:
        base_img = camera_image.copy().convert("RGBA")
    else:
        base_img = PILImage.new("RGBA", (img_w, img_h), (50, 50, 50, 255))

    # Cycle through Kingfisher Normal, Dark, and Light RGB values based on rule count
    rgb_color = [(0, 161, 154), (0, 114, 110), (229, 245, 245)][index % 3]
    
    overlay = PILImage.new("RGBA", base_img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    
    draw.polygon(vertices, fill=rgb_color + (76,))
    draw.polygon(vertices, outline=rgb_color + (255,), width=3)
    
    for (x, y) in vertices:
        r = 6
        draw.ellipse((x - r, y - r, x + r, y + r), fill=(255, 255, 0, 255), outline=(0, 0, 0, 255), width=1)
    
    final_img = PILImage.alpha_composite(base_img, overlay).convert("RGB")
    buf = BytesIO()
    final_img.save(buf, format="PNG")
    buf.seek(0)

    img = PILImage.open(buf)
    ratio = THUMB_H / img.height
    img_resized = img.resize((int(img.width * ratio), THUMB_H), PILImage.Resampling.LANCZOS)
    
    img_buf = BytesIO()
    img_resized.save(img_buf, format="PNG")
    img_buf.seek(0)
    
    return OpenpyxlImage(img_buf)

def create_new_workbook(batch_num):
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Camera Analytics"
    ws_main.views.sheetView[0].showGridLines = True

    ws_missed = wb.create_sheet(title="Missed Cameras")
    ws_missed.views.sheetView[0].showGridLines = True

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="1A1D27")
    failed_fill   = PatternFill("solid", start_color="A61C1C") 
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_main.merge_cells("A1:I1")
    ws_main["A1"] = f"Intelligent Analytics Master Report - Part {batch_num}"
    ws_main["A1"].font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws_main["A1"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A1"].alignment = center_align
    ws_main.row_dimensions[1].height = 36

    ws_main.merge_cells("A2:I2")
    ws_main["A2"] = f"Batch Processing Timeline Context: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_main["A2"].font      = Font(name="Arial", size=10, color="888888")
    ws_main["A2"].fill      = PatternFill("solid", start_color="0F1117")
    ws_main["A2"].alignment = center_align
    ws_main.row_dimensions[2].height = 20

    main_headers = [
        "Client Name", "Location", "Live Unit Serial", "Camera Position", 
        "Rule Name", "Rule Type", "Duration (s)", "Target Detection", "Rule Visual Overlay Thumbnail"
    ]
    for col, h in enumerate(main_headers, 1):
        cell = ws_main.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        cell.border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                             top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
    ws_main.row_dimensions[3].height = 22

    main_widths = [22, 24, 22, 16, 16, 20, 14, 20, 110]  
    for i, w in enumerate(main_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = w

    ws_missed.merge_cells("A1:G1")
    ws_missed["A1"] = "Analytics Fetch Exception Audit Log"
    ws_missed["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_missed["A1"].fill      = PatternFill("solid", start_color="330000") 
    ws_missed["A1"].alignment = center_align
    ws_missed.row_dimensions[1].height = 36

    missed_headers = ["Timestamp", "Client Name", "Location", "Live Unit Serial", "Camera Port", "Target Endpoint", "Failure Reason"]
    for col, h in enumerate(missed_headers, 1):
        cell = ws_missed.cell(row=2, column=col, value=h)
        cell.font = header_font; cell.fill = failed_fill; cell.alignment = center_align
    ws_missed.row_dimensions[2].height = 22

    missed_widths = [22, 22, 24, 22, 15, 65, 45]
    for i, w in enumerate(missed_widths, 1):
        ws_missed.column_dimensions[get_column_letter(i)].width = w

    ws_main.freeze_panes = "A4"
    ws_missed.freeze_panes = "A3"
    
    return wb, ws_main, ws_missed

# --- 3. DYNAMIC MODE & SELECTION RUNTIME PROMPTS ---
print("=========================================")
print("   HIKVISION ANALYTICS REPORT GENERATOR  ")
print("=========================================")
print("1. Process a batch of cameras via CSV file")
print("2. Test a single camera / IP address")
print("-----------------------------------------")
run_mode = input("Select an option (1 or 2) [Default 1]: ").strip()

camera_rows = []
base_filename = ""
DATE_SUFFIX = datetime.now().strftime("%Y%m%d")

if run_mode == "2":
    print("\n[ Single Camera Mode Selected ]")
    single_ip = input("Enter the Camera IP Address: ").strip()
    while not single_ip:
        single_ip = input("[!] IP Address cannot be blank. Enter Camera IP: ").strip()
        
    camera_rows.append({
        "CLIENT_NM": input("Enter Client Name (Optional): ").strip() or "Single Test",
        "LOCATION_NM": input("Enter Location Name (Optional): ").strip() or "Diagnostic",
        "LIVE_UNIT_SERIAL_NM": input("Enter Unit Serial (Optional): ").strip() or "N/A",
        "IP": single_ip
    })
    base_filename = f"Hik_Single_Test_{single_ip.replace('.', '_')}"
else:
    print("\n[ CSV Batch Mode Selected ]")
    print("Select your input CSV file via the window popup...")
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    selected_file = filedialog.askopenfilename(
        initialdir=downloads_path,
        title="Select Camera Layout CSV File",
        filetypes=(("CSV Files", "*.csv"), ("All Files", "*.*"))
    )
    
    if selected_file:
        csv_input_path = Path(selected_file)
    else:
        print("[-] Window closed without selection. Falling back to manual name lookups.")
        csv_input_path = downloads_path / (input("Enter the input CSV file name: ").strip() or "fullTesting.csv")

    if not csv_input_path.exists():
        print(f"[!] Error: Missing input file profile target: {csv_input_path}")
        exit()

    print(f"Reading target camera layouts from: {csv_input_path}...")
    with open(csv_input_path, mode='r', encoding='utf-8-sig') as f:
        camera_rows = list(csv.DictReader(f))
            
    custom_tag = input("Enter an optional custom tag/job name (or press Enter to use CSV name): ").strip()
    base_filename = custom_tag if custom_tag else csv_input_path.stem

REPORT_PART_FILE_TEMPLATE = f"{base_filename}_Part_{{batch_idx}}_{DATE_SUFFIX}.xlsx"
FINAL_ZIP_ARCHIVE_NAME = f"{base_filename}_Archive_{DATE_SUFFIX}.zip"

# --- 4. EXECUTION ENGINE INIT ---
generated_files = []
batch_index = 1
current_main_row = 4
current_missed_row = 3
rows_written_in_current_file = 0

wb, ws_main, ws_missed = create_new_workbook(batch_index)

normal_font = Font(name="Arial", size=10)
bold_font   = Font(name="Arial", bold=True, size=10)

white_normal_font = Font(name="Arial", size=10, color="FFFFFF")
white_bold_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")

left_align  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
center_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                     top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

# Connection Pool Initialization
session = requests.Session()
auth_strategies = [HTTPDigestAuth(USER, PASSWORD), HTTPBasicAuth(USER, PASSWORD)]

# --- 5. CORE RUNTIME ENGINE LOOP ---
for row_data in camera_rows:
    client_name = row_data.get("CLIENT_NM", "").strip()
    location = row_data.get("LOCATION_NM", "").strip()
    serial = row_data.get("LIVE_UNIT_SERIAL_NM", "").strip()
    ip = row_data.get("IP", "").strip()
    
    if not ip:
        continue
        
    print(f"\n#################################################################")
    print(f" INGESTING HOST NODE: {client_name} - {location} ({ip})")
    print(f"#################################################################")

    for cam in CAMERA_CONFIGS:
        port = cam["port"]
        position = cam["position"]
        bg_color = cam["color"]
        row_fill = PatternFill("solid", start_color=bg_color)
        
        # Automatically use white text if the background is Kingfisher Dark
        if bg_color == "00726E":  
            active_normal_font = white_normal_font
            active_bold_font   = white_bold_font
        else:
            active_normal_font = normal_font
            active_bold_font   = bold_font
            
        print(f" -> Testing {position} Interface Port: {port}...")

        # A. FETCH STREAM BACKGROUND SNAPSHOT
        camera_image = fetch_camera_snapshot(session, ip, port, auth_strategies)
        if camera_image is None:
            print("    [!] Warning: Failed to fetch stream snapshot. Defaulting to 1920x1080 canvas.")
            img_w, img_h = 1920, 1080
        else:
            img_w, img_h = camera_image.size

        # B. FETCH BEHAVIOR RULES PAYLOAD
        xml_data, rule_url_used = fetch_active_rules(session, ip, port, auth_strategies)

        if xml_data is None:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws_missed.append([timestamp, client_name, location, serial, f"{position} ({port})", f"http://{ip}:{port}/ISAPI/Intelligent/channels/.../behaviorRule/1", f"No active analytic perimeters found"])
            for col in range(1, 8):
                ws_missed.cell(row=current_missed_row, column=col).border = thin_border
                ws_missed.cell(row=current_missed_row, column=col).font = normal_font
            current_missed_row += 1
            continue

        # C. PARSE RULE PAYLOAD ELEMENTS
        rules = parse_analytics_xml(xml_data, img_w, img_h)
        
        if not rules:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws_missed.append([timestamp, client_name, location, serial, f"{position} ({port})", rule_url_used, "Missing explicit RuleInfo elements or Parsing Error"])
            for col in range(1, 8):
                ws_missed.cell(row=current_missed_row, column=col).border = thin_border
            current_missed_row += 1
            continue

        # D. ROW GENERATION FOR DETECTED ANALYTICS
        for index, rule in enumerate(rules):
            if rows_written_in_current_file >= MAX_ROWS_PER_FILE:
                file_path = output_dir / REPORT_PART_FILE_TEMPLATE.format(batch_idx=batch_index)
                wb.save(file_path)
                generated_files.append(file_path)
                print(f"    [#] Saved workbook partition part limit reached: {file_path}")
                
                batch_index += 1
                wb, ws_main, ws_missed = create_new_workbook(batch_index)
                current_main_row = 4
                current_missed_row = 3
                rows_written_in_current_file = 0

            def styled_main(col, value, font=active_normal_font, align=left_align):
                c = ws_main.cell(row=current_main_row, column=col, value=value)
                c.font = font; c.fill = row_fill; c.alignment = align; c.border = thin_border
                return c

            styled_main(1, client_name, align=left_align)
            styled_main(2, location, align=left_align)
            styled_main(3, serial, align=center_align)
            styled_main(4, position, font=active_bold_font, align=center_align)
            styled_main(5, rule["name"], align=left_align)
            styled_main(6, rule["type"], align=left_align)
            
            dur = rule["duration"]
            styled_main(7, int(dur) if dur.isdigit() else dur, align=center_align)
            styled_main(8, rule["target"], align=center_align)

            thumb_cell = ws_main.cell(row=current_main_row, column=9, value="")
            thumb_cell.fill = row_fill; thumb_cell.border = thin_border

            # Generate and anchor overlay image
            try:
                xl_img = render_overlay_image(camera_image, rule["vertices"], index, img_w, img_h)
                
                start_marker = AnchorMarker(col=8, colOff=0, row=current_main_row-1, rowOff=0)
                end_marker   = AnchorMarker(col=9, colOff=0, row=current_main_row, rowOff=0)
                xl_img.anchor = TwoCellAnchor(editAs="oneCell", _from=start_marker, to=end_marker)
                
                ws_main.add_image(xl_img)
                print(f"    [+] Logged rule details for Rule Name: {rule['name']} (Duration: {dur}s)")
            except Exception as img_err:
                thumb_cell.value = f"(Image failed: {img_err})"

            ws_main.row_dimensions[current_main_row].height = ROW_H
            current_main_row += 1
            rows_written_in_current_file += 1

# --- 6. EXPORT & COMPRESSION ENGINE ---
file_path = output_dir / REPORT_PART_FILE_TEMPLATE.format(batch_idx=batch_index)

if batch_index == 1 and rows_written_in_current_file < MAX_ROWS_PER_FILE:
    file_path = downloads_path / f"{base_filename}_{DATE_SUFFIX}.xlsx"

wb.save(file_path)
generated_files.append(file_path)

if len(generated_files) > 1:
    zip_output_path = downloads_path / FINAL_ZIP_ARCHIVE_NAME
    print(f"\n[+] Multiple files generated ({len(generated_files)} parts). Packaging into ZIP container...")
    
    with zipfile.ZipFile(zip_output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in generated_files:
            zipf.write(file, arcname=file.name)
            os.remove(file)
            
    print(f"[+] Success! Consolidated archives compiled into:\n    {zip_output_path}")
    try:
        os.rmdir(output_dir)
    except Exception:
        pass
else:
    if generated_files and run_mode != "2":
        try:
            final_moved_path = downloads_path / generated_files[0].name
            os.rename(generated_files[0], final_moved_path)
            os.rmdir(output_dir)
            print(f"\n[+] Success! Output saved cleanly to:\n    {final_moved_path}")
        except Exception:
            print(f"\n[+] Success! Output saved cleanly to:\n    {generated_files[0]}")
    elif run_mode == "2":
        print(f"\n[+] Success! Diagnostic test completed. Output saved to downloads folder.")

# --- 7. TIME BENCHMARK RUN ---
elapsed_seconds = time.time() - script_start_time
minutes = int(elapsed_seconds // 60)
seconds = int(elapsed_seconds % 60)

print(f"\n=================================================================")
print(f" HIKVISION BATCH PROCESSING COMPLETED SUCCESSFULLY")
print(f"=================================================================")
print(f" -> Total Duration: {minutes}m {seconds}s ({round(elapsed_seconds, 2)} total seconds)")
print(f"=================================================================\n")