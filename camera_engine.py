"""Camera analytics engine: fetching, parsing, rendering, and report generation.

No input()/print()/exit() here -- this module is a plain library used by both
the CLI (combined.py) and the GUI (audit_gui.py). Callers pass in credentials and
a log/progress callback instead of relying on module-level globals or stdout.
"""

import concurrent.futures
import time
import xml.etree.ElementTree as ET
from datetime import datetime
import os
from io import BytesIO
from pathlib import Path

import requests
import urllib3
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from PIL import Image as PILImage, ImageDraw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

STRICT_TIMEOUT = (3.05, 5.0)
MAX_WORKER_THREADS = 15

THUMB_H = 450
ROW_H = 340

CAMERA_CONFIGS = [
    {"port": "5010", "position": "CENTER", "color": "E5F5F5"},  # LVT Light
    {"port": "5015", "position": "LEFT",   "color": "00A19A"},  # LVT Normal
    {"port": "5020", "position": "RIGHT",  "color": "00726E"}   # LVT Dark
]


def default_output_dir():
    downloads_path = Path.home() / "Downloads"
    output_dir = downloads_path / "Camera_Reports_Master"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def classify_manufacturer(raw_value):
    """Returns 'AXIS', 'HIKVISION', 'MIXED', or None if the value doesn't clearly
    match any of those. Never silently defaults -- an unrecognized value (blank,
    misspelled, wrong CSV column mapped in) must be caught by the caller instead
    of routed to the wrong API. "LVT" is a rebranded/OEM camera line running
    Hikvision firmware -- same ISAPI endpoints, same credentials -- so it's
    treated as HIKVISION here. "MIXED" means the physical unit's three camera
    positions aren't all the same vendor (e.g. two Hikvision, one Axis) -- see
    process_camera_row's MIXED branch for how that's actually resolved per port."""
    value = (raw_value or "").strip().lower()
    if "mixed" in value or "both" in value:
        return "MIXED"
    if "axis" in value:
        return "AXIS"
    if "hik" in value or "lvt" in value:
        return "HIKVISION"
    return None


def dedupe_camera_rows(rows):
    """Collapse rows sharing the same IP into a single device row.

    If an IP appears multiple times with more than one recognized vendor, the
    resulting row is marked as MANUFACTURER='mixed' so the engine can probe each
    port against both Axis and Hikvision APIs.
    """
    grouped = {}
    order = []
    passthrough = []
    for row in rows:
        ip = (row.get("IP", "") or "").strip()
        if not ip:
            passthrough.append(row)
            continue
        if ip not in grouped:
            grouped[ip] = []
            order.append(ip)
        grouped[ip].append(row)

    deduped = []
    for ip in order:
        group = grouped[ip]
        if len(group) == 1:
            deduped.append(group[0])
            continue

        classes = {classify_manufacturer(r.get("MANUFACTURER", "")) for r in group}
        classes.discard(None)
        base = dict(group[0])
        if len(classes) > 1:
            base["MANUFACTURER"] = "mixed"
        deduped.append(base)

    return deduped + passthrough


def _extract_hik_channel_id(behavior_rule_url):
    """Pulls the channel id back out of a successful Hikvision behaviorRule URL,
    e.g. '.../channels/2/behaviorRule/1' -> '2'. Returns None for anything that
    doesn't match that shape (an Axis URL, or the bracketed placeholder URL used
    on a failed fetch)."""
    marker = "/channels/"
    if marker not in behavior_rule_url or "/behaviorRule/" not in behavior_rule_url:
        return None
    try:
        candidate = behavior_rule_url.split(marker, 1)[1].split("/", 1)[0]
    except IndexError:
        return None
    return candidate if candidate.isdigit() else None


# --- OBJECT-ORIENTED ARCHITECTURE (OOP) ---

class CameraHandler:
    """Base class for camera manufacturers."""
    def __init__(self, ip, user, password):
        self.ip = ip
        self.auth_strategies = [HTTPDigestAuth(user, password), HTTPBasicAuth(user, password)]

    def fetch_snapshot(self, session, port):
        raise NotImplementedError

    def fetch_analytics(self, session, port):
        raise NotImplementedError

    def reposition_for_scene(self, session, port, channel_id, rule_index):
        """Optional hook for camera types that need physical repositioning before a
        snapshot will match an analytics rule's coordinate space (e.g. a PTZ camera
        that binds a rule to a saved pan/tilt/zoom preset). No-op by default --
        fixed cameras (and Axis, until proven otherwise) don't need this."""
        return False

    def parse_analytics(self, data, img_w, img_h):
        raise NotImplementedError


class HikvisionHandler(CameraHandler):
    def __init__(self, ip, user, password):
        super().__init__(ip, user, password)
        self.fallback_dim = (1920, 1080)
        self.is_thermal_device = False

    # PTZ-capable Hikvision units bind an analytics rule's zone coordinates to a
    # saved pan/tilt/zoom position (a "scenePtz"): the camera's own web UI issues a
    # PUT to scenePtz/<id>/goto and waits before showing/editing a rule, because the
    # rule's coordinates are only meaningful once the camera is actually parked at
    # that position -- otherwise a live snapshot and the rule's coordinate space are
    # simply looking at two different framings. This was reproduced and confirmed via
    # a real HAR capture of the web UI against 10.23.27.20. The 6.5s wait mirrors the
    # ~6.3s the web UI itself waited (polling lockPTZ) between issuing goto and using
    # the rule/live view again -- an observed real value, not a guess.
    PTZ_SETTLE_WAIT_SECONDS = 6.5

    def reposition_for_scene(self, session, port, channel_id, rule_index):
        goto_url = f"http://{self.ip}:{port}/ISAPI/Intelligent/channels/{channel_id}/scenePtz/{rule_index}/goto"
        for auth in self.auth_strategies:
            try:
                response = session.put(goto_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
                if response.status_code == 200:
                    time.sleep(self.PTZ_SETTLE_WAIT_SECONDS)
                    return True
            except requests.exceptions.RequestException:
                continue
        return False

    def fetch_snapshot(self, session, port):
        # ALWAYS attempt to pull the Thermal channel (201) first.
        # If the camera does not physically possess a thermal layout, 
        # it will fail and fall back to the Optical channel (101).
        channels_to_try = ["201", "101"]
        
        last_err = "All authentication attempts failed"
        saw_401 = False
        saw_other = False

        for channel in channels_to_try:
            img_url = f"http://{self.ip}:{port}/ISAPI/Streaming/channels/{channel}/picture"
            for auth in self.auth_strategies:
                try:
                    with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, stream=True, verify=False) as response:
                        if response.status_code == 200:
                            if channel == "201":
                                self.is_thermal_device = True
                            return PILImage.open(BytesIO(response.content)).copy(), img_url, None, False
                        if response.status_code == 401:
                            saw_401 = True
                        else:
                            saw_other = True
                        last_err = f"Channel {channel} HTTP {response.status_code}"
                except requests.exceptions.RequestException as e:
                    saw_other = True
                    last_err = f"Channel {channel} Network Error: {type(e).__name__}"
                    continue

        auth_rejected = saw_401 and not saw_other
        fallback_url = f"http://{self.ip}:{port}/ISAPI/Streaming/channels/101/picture"
        return None, fallback_url, last_err, auth_rejected

    def fetch_analytics(self, session, port):
        last_status = None
        last_snippet = None
        saw_401 = False
        saw_other = False
        saw_ok = False
        for channel_id in [1, 2]:
            rule_url = f"http://{self.ip}:{port}/ISAPI/Intelligent/channels/{channel_id}/behaviorRule/1"
            for auth in self.auth_strategies:
                try:
                    response = session.get(rule_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
                    last_status = response.status_code
                    if response.status_code == 200:
                        temp_xml = response.text
                        if temp_xml and ("positionX" in temp_xml or "RegionCoordinates" in temp_xml):
                            return temp_xml, rule_url, None, False
                        # A clean 200 with no coordinate tags means the camera answered
                        # fine, it just has no analytic zones configured on this channel.
                        saw_ok = True
                        last_snippet = (temp_xml[:200] + "...") if temp_xml and len(temp_xml) > 200 else (temp_xml or "(empty body)")
                    elif response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                except requests.exceptions.RequestException as e:
                    saw_other = True
                    last_snippet = f"{type(e).__name__}: {e}"
                    continue
        auth_rejected = saw_401 and not saw_other and not saw_ok
        if last_status is None:
            diag = f"No response from any channel/auth combo (last error: {last_snippet})"
        else:
            diag = f"Last HTTP {last_status} response did not contain expected rule tags. Body snippet: {last_snippet}"

        url = f"http://{self.ip}:{port}/ISAPI/Intelligent/channels/[1,2]/behaviorRule/1"

        # Thermal/bi-spectrum units (e.g. DS-2TD4237-7/V2) are intentionally NOT
        # special-cased here. Verified against real device traffic (see
        # test_thermal_channels.py): these cameras serve their VCA rules through the
        # same behaviorRule endpoint on channels 1/2 and a snapshot on channel 201, so
        # they flow through the identical path as any other Hikvision unit above.
        #
        # NO_RULES signals that the camera responded cleanly (HTTP 200) but has no
        # analytic zones configured -- optical or thermal alike. The caller treats this
        # as a normal empty result: it still renders the snapshot and does NOT log the
        # unit to the Missed tab. Only genuine failures (auth/network/error status)
        # below produce a Missed entry.
        if saw_ok:
            return None, url, f"NO_RULES: No active analytic perimeters configured on this device | {diag}", auth_rejected
        return None, url, f"No active analytic perimeters found | {diag}", auth_rejected

    def parse_analytics(self, xml_data, img_w, img_h):
        namespaces = {'ns': 'http://www.std-cgi.com/ver20/XMLSchema'}
        try:
            root = ET.fromstring(xml_data)
            rules = root.findall('.//ns:RuleInfo', namespaces)
        except Exception:
            rules = []

        if not rules:
            return [{"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]

        parsed_rules = []
        for rule in rules:
            rule_name = rule.find('ns:ruleName', namespaces).text

            event_type_raw = rule.find('ns:eventType', namespaces)
            event_type = event_type_raw.text if event_type_raw is not None else "Unknown"
            if "field" in event_type.lower() or "intrusion" in event_type.lower():
                event_type = "Intrusion Detection"
            elif "line" in event_type.lower() or "cross" in event_type.lower():
                event_type = "Line Crossing"

            duration_val = "0"
            possible_tags = ['.//ns:FieldDetectionParam/ns:durationTime', './/ns:FieldDetectionParam/ns:timeDuration', './/ns:LineCrossing/ns:duration', './/ns:Intrusion/ns:duration', './/ns:durationTime', './/ns:timeDuration', './/ns:duration', './/ns:alarmDelay']
            found_values = [rule.find(tag, namespaces).text.strip() for tag in possible_tags if rule.find(tag, namespaces) is not None and rule.find(tag, namespaces).text]
            if found_values:
                valid_seconds = [v for v in found_values if v != "100" and v != "0"]
                duration_val = valid_seconds[0] if valid_seconds else found_values[0]

            target_node = rule.find('.//ns:FieldDetectionParam/ns:detectionTarget', namespaces)
            if target_node is None:
                target_node = rule.find('.//ns:detectionTarget', namespaces) or rule.find('.//ns:TargetType', namespaces)
            target_detection = target_node.text.capitalize() if (target_node is not None and target_node.text) else "All Targets"

            # Line-crossing direction. Hik lineDetection stores directionSensitivity
            # (left-right / right-left / any) inside LineDetectionParam. Map to the
            # neutral Axis-style values the renderer expects (leftToRight/rightToLeft/
            # any) using the same mapping vendor_adapter applies on the writer's read
            # path -- kept in sync here. The Hik parser vertically flips vertices to
            # match snapshot orientation; the arrow's rendered side was VERIFIED against
            # a live device (10.23.5.188 ch2 "TESTZONE" = right-left) on 2026-08-03 --
            # the audit overlay's arrow matched the camera web UI's direction.
            direction = None
            if event_type == "Line Crossing":
                ds_node = rule.find('.//ns:LineDetectionParam/ns:directionSensitivity', namespaces)
                ds_val = ds_node.text.strip() if (ds_node is not None and ds_node.text) else None
                direction = {"left-right": "leftToRight", "right-left": "rightToLeft",
                             "any": "any"}.get(ds_val)

            region_lists = rule.findall('.//ns:RegionCoordinatesList', namespaces)
            polygons = []
            for region in region_lists:
                vertices = []
                for coord in region.findall('ns:RegionCoordinates', namespaces):
                    raw_x = float(coord.find('ns:positionX', namespaces).text)
                    raw_y = float(coord.find('ns:positionY', namespaces).text)

                    # Hikvision positionX/positionY are in a 0..1000 coordinate
                    # space where (0,0) is the top-left. Map directly to pixel
                    # coordinates and clamp to image bounds. Previous behavior
                    # inverted both axes which produced mirrored overlays vs the
                    # camera web UI.
                    pixel_x = int((raw_x / 1000.0) * img_w)
                    pixel_y = int((raw_y / 1000.0) * img_h)
                    # This camera's rule coordinates require a vertical flip to match
                    # the live snapshot orientation observed in test_variant_flip_y.
                    pixel_y = img_h - 1 - pixel_y
                    pixel_x = max(0, min(img_w - 1, pixel_x))
                    pixel_y = max(0, min(img_h - 1, pixel_y))
                    vertices.append((pixel_x, pixel_y))

                if vertices:
                    polygons.append(vertices)

            if not polygons:
                continue

            # Min/max object-size boxes from SizeFilter. positionX/Y/width/height share
            # the 0..1000 grid as the region, so apply the same vertical flip: the box
            # spans posY..posY+h in raw coords, which flips to top = h-(posY+h), bottom = h-posY.
            size_boxes = []
            for box_tag, label in (("MinObjectSize", "min size"), ("MaxObjectSize", "max size")):
                box = rule.find(f'.//ns:SizeFilter/ns:{box_tag}', namespaces)
                if box is None:
                    continue
                try:
                    bx = float(box.find('ns:positionX', namespaces).text)
                    by = float(box.find('ns:positionY', namespaces).text)
                    bw = float(box.find('ns:width', namespaces).text)
                    bh = float(box.find('ns:height', namespaces).text)
                except (AttributeError, TypeError, ValueError):
                    continue
                if bw <= 0 or bh <= 0:
                    continue  # default/empty SizeFilter -- nothing to draw
                x0 = int((bx / 1000.0) * img_w)
                x1 = int(((bx + bw) / 1000.0) * img_w)
                y0 = int(img_h - 1 - ((by + bh) / 1000.0) * img_h)
                y1 = int(img_h - 1 - (by / 1000.0) * img_h)
                size_boxes.append({"box": (x0, y0, x1, y1), "label": label})

            for zone_index, polygon in enumerate(polygons, start=1):
                zone_name = rule_name if len(polygons) == 1 else f"{rule_name} [Zone {zone_index}]"
                parsed_rules.append({
                    "is_placeholder": False,
                    "name": zone_name,
                    "type": event_type,
                    "duration": duration_val,
                    "target": target_detection,
                    "vertices": polygon,
                    "size_boxes": size_boxes,
                    "bars": [],
                    "direction": direction,
                })

        # Deduplicate parsed rules that are exactly identical (same name/type/target/duration/vertices)
        def normalize_vertices_key(verts):
            if not verts:
                return ()
            if isinstance(verts[0], (list, tuple)) and verts[0] and isinstance(verts[0][0], (list, tuple)):
                return tuple(tuple((int(x), int(y)) for x, y in polygon) for polygon in verts)
            return tuple((int(x), int(y)) for x, y in verts)

        seen = set()
        deduped = []
        for r in parsed_rules:
            verts_key = normalize_vertices_key(r.get("vertices", []))
            key = (r.get("name"), r.get("type"), r.get("target"), r.get("duration"), verts_key)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)

        if deduped:
            return deduped
        return [{"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]


class AxisHandler(CameraHandler):
    def __init__(self, ip, user, password):
        super().__init__(ip, user, password)
        self.fallback_dim = (1280, 720)

    def fetch_snapshot(self, session, port):
        img_url = f"http://{self.ip}:{port}/axis-cgi/jpg/image.cgi?resolution=1280x720"
        last_err = "All authentication attempts failed"
        saw_401 = False
        saw_other = False
        for auth in self.auth_strategies:
            # --- DUAL METHOD FALLBACK: TRY GET FIRST ---
            try:
                with session.get(img_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False) as response:
                    if response.status_code == 200:
                        return PILImage.open(BytesIO(response.content)).copy(), img_url, None, False
                    if response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                    last_err = f"GET HTTP {response.status_code}"
            except requests.exceptions.RequestException as e:
                saw_other = True
                last_err = f"GET Error: {type(e).__name__}"

            # --- DUAL METHOD FALLBACK: RETRY VIA POST IF GET REJECTED/NON-200 ---
            try:
                with session.post(img_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False) as response:
                    if response.status_code == 200:
                        return PILImage.open(BytesIO(response.content)).copy(), img_url, None, False
                    if response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                    last_err = f"POST HTTP {response.status_code}"
            except requests.exceptions.RequestException as e:
                saw_other = True
                last_err = f"POST Error: {type(e).__name__}"

        auth_rejected = saw_401 and not saw_other
        return None, img_url, last_err, auth_rejected

    def _detect_active_analytics_app(self, session, port):
        """Called only when the generic Object Analytics endpoint 404s. Checks the
        camera's installed-application list to identify what analytics engine (if
        any) is actually running -- e.g. AXIS Perimeter Defender on fixed thermal
        units, which is a separate ACAP app with no relation to Object Analytics."""
        list_url = f"http://{self.ip}:{port}/axis-cgi/applications/list.cgi"
        for auth in self.auth_strategies:
            try:
                response = session.get(list_url, auth=auth, timeout=STRICT_TIMEOUT, verify=False)
                if response.status_code == 200:
                    try:
                        root = ET.fromstring(response.text)
                    except ET.ParseError:
                        return None
                    return [app.get("NiceName", app.get("Name", "Unknown App"))
                            for app in root.findall(".//application")
                            if app.get("Status") == "Running"]
            except requests.exceptions.RequestException:
                continue
        return None

    # AXIS Perimeter Defender (common on fixed THERMAL units, e.g. Q1971-E) is a
    # separate ACAP with its own v2 REST API -- NOT the Object Analytics control.cgi
    # endpoint. Confirmed against a real packet capture of the camera web UI: the
    # configured alarm zones are published on this multipart metadata stream as
    # <ALERT_ZONE_LIST> with pixel-space POINT coordinates in the OSD_SIZE frame.
    PD_METADATA_PATH = "/local/AXISPerimeterDefender/v2/metadata/liveStream"

    def _fetch_perimeter_defender(self, session, port):
        """Read a single metadata frame from the Perimeter Defender live stream.

        The endpoint is a multipart/x-mixed-replace stream that emits an XML <NODE>
        frame ~9x/second; we read only until the first complete frame arrives, then
        drop the connection. Returns (xml_frame, url, err)."""
        url = f"http://{self.ip}:{port}{self.PD_METADATA_PATH}"
        for auth in self.auth_strategies:
            try:
                with session.get(url, auth=auth, timeout=STRICT_TIMEOUT, stream=True, verify=False) as resp:
                    if resp.status_code != 200:
                        continue
                    buf = ""
                    for chunk in resp.iter_content(chunk_size=2048):
                        if not chunk:
                            continue
                        buf += chunk.decode("utf-8", errors="replace")
                        end = buf.find("</NODE>")
                        if end != -1:
                            start = buf.find("<NODE")
                            if start != -1:
                                return buf[start:end + len("</NODE>")], url, None
                        if len(buf) > 131072:  # safety cap -- never read the stream forever
                            break
            except requests.exceptions.RequestException:
                continue
        return None, url, "Perimeter Defender metadata stream returned no zone frame"

    def _parse_perimeter_defender(self, xml_frame, img_w, img_h):
        """Turn one Perimeter Defender metadata frame into rule dicts. POINT
        coordinates are already pixels in the OSD_SIZE space, so we scale by
        snapshot_size / OSD_SIZE rather than doing a normalized-coordinate map."""
        try:
            root = ET.fromstring(xml_frame)
        except ET.ParseError:
            return [{"is_placeholder": True, "name": "Perimeter Defender (unparseable metadata)", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]

        try:
            ref_w, ref_h = (int(v) for v in root.get("OSD_SIZE", "384x288").lower().split("x"))
        except (ValueError, AttributeError):
            ref_w, ref_h = 384, 288
        if ref_w <= 0 or ref_h <= 0:
            ref_w, ref_h = 384, 288

        rules = []
        for zone in root.findall(".//ALERT_ZONE_LIST/ZONE"):
            points = sorted(zone.findall("POINT"), key=lambda p: int(p.get("NUMBER", "0") or "0"))
            vertices = []
            for p in points:
                try:
                    raw_x = float(p.get("X2D", "0"))
                    raw_y = float(p.get("Y2D", "0"))
                except ValueError:
                    continue
                px = max(0, min(img_w - 1, int(raw_x / ref_w * img_w)))
                py = max(0, min(img_h - 1, int(raw_y / ref_h * img_h)))
                vertices.append((px, py))
            if not vertices:
                continue
            zone_name = zone.get("NAME", "zone")
            rules.append({
                "is_placeholder": False,
                "name": f"Perimeter Defender: {zone_name}",
                # This frame carries zone geometry but not the scenario's configured
                # type/target/duration (those live in the binary context.knp). Label
                # generically rather than inventing specifics we didn't observe.
                "type": "Perimeter Defender",
                "target": "Human / Vehicle",
                "duration": "N/A",
                "vertices": vertices,
            })

        if not rules:
            return [{"is_placeholder": True, "name": "Perimeter Defender: No Zones Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]
        return rules

    def fetch_analytics(self, session, port):
        control_url = f"http://{self.ip}:{port}/local/objectanalytics/control.cgi"
        payload = {"apiVersion": "1.2", "method": "getConfiguration"}
        last_err = "All authentication attempts failed"
        saw_404 = False
        saw_401 = False
        saw_other = False

        for auth in self.auth_strategies:
            try:
                response = session.post(control_url, auth=auth, json=payload, timeout=STRICT_TIMEOUT, verify=False)
                if response.status_code == 200:
                    json_data = response.json()
                    # Axis's local CGI APIs return HTTP 200 with an "error" body on failure
                    # (e.g. Object Analytics app not running/licensed, unsupported apiVersion)
                    # instead of a non-200 status. Without this check that error was silently
                    # treated as "zero scenarios configured".
                    if isinstance(json_data, dict) and "error" in json_data:
                        err = json_data["error"]
                        if isinstance(err, dict):
                            last_err = f"API error {err.get('code', '?')}: {err.get('message', json_data)}"
                        else:
                            last_err = f"API error: {err}"
                        saw_other = True
                        continue
                    return json_data, control_url, None, False
                else:
                    if response.status_code == 404:
                        saw_404 = True
                    elif response.status_code == 401:
                        saw_401 = True
                    else:
                        saw_other = True
                    last_err = f"HTTP Error Status {response.status_code}: {response.text[:200]}"
            except Exception as e:
                saw_other = True
                last_err = f"{type(e).__name__}: {e}"
                continue

        auth_rejected = saw_401 and not saw_other and not saw_404
        if saw_404:
            running_apps = self._detect_active_analytics_app(session, port)
            if running_apps and any("perimeter" in app.lower() for app in running_apps):
                # Fixed thermal units typically run Perimeter Defender instead of
                # Object Analytics -- pull the zones from its metadata stream.
                pd_xml, pd_url, _pd_err = self._fetch_perimeter_defender(session, port)
                if pd_xml is not None:
                    return {"__perimeter_defender__": pd_xml}, pd_url, None, False
            if running_apps:
                last_err = (f"SPECIAL CASE: This device has no AXIS Object Analytics app -- "
                            f"active analytics app instead: {', '.join(running_apps)}. "
                            f"Requires a separate integration, not a code failure.")
        return None, control_url, last_err, auth_rejected

    def parse_analytics(self, json_data, img_w, img_h):
        # Perimeter Defender payloads come through as a tagged dict carrying one
        # metadata XML frame instead of the Object Analytics JSON shape.
        if isinstance(json_data, dict) and "__perimeter_defender__" in json_data:
            return self._parse_perimeter_defender(json_data["__perimeter_defender__"], img_w, img_h)

        scenarios = json_data.get("data", {}).get("scenarios", [])
        if not scenarios:
            return [{"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]

        # Perspective calibration bars live at the top level and scenarios reference them
        # by id; index them so each rule can carry its bars into the overlay.
        persp_defs = {p.get("id"): p for p in json_data.get("data", {}).get("perspectives", [])}

        def _aoa_to_px(x, y):
            return (int(((float(x) + 1.0) / 2.0) * img_w), int(((1.0 - float(y)) / 2.0) * img_h))

        parsed_rules = []
        for scenario in scenarios:
            if scenario.get("is_placeholder"):
                parsed_rules.append({"is_placeholder": True, "name": "No Scenarios Configured", "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []})
                continue

            rule_name = scenario.get("name", f"Scenario {scenario.get('id')}")
            loiter_time = ""
            triggers = scenario.get("triggers", [])

            for trigger_obj in triggers:
                for condition in trigger_obj.get("conditions", []):
                    if condition.get("type") == "individualTimeInArea":
                        for d in condition.get("data", []):
                            if d.get("time"):
                                loiter_time = str(d["time"])
                                break

            if not loiter_time:
                for f_obj in scenario.get("filters", []):
                    if f_obj.get("type") == "timeShortLivedLimit":
                        loiter_time = str(f_obj.get("time", ""))
                        break

            duration_val = loiter_time if loiter_time else "0"
            classes = [oc.get("type", "") for oc in scenario.get("objectClassifications", []) if oc.get("type")]
            target_detection = ", ".join(c.capitalize() + " Detection" for c in classes) if classes else "Any Detection"

            rule_type_base = scenario.get("type", "")
            if triggers and not rule_type_base: rule_type_base = triggers[0].get("type", "Unknown")
            rule_type = "Intrusion Detection" if "area" in rule_type_base.lower() else rule_type_base.capitalize()

            vertices = []
            if triggers:
                raw_vertices = triggers[0].get("vertices", [])
                for pt in raw_vertices:
                    vertices.append(_aoa_to_px(pt[0], pt[1]))

            # Perspective calibration bars for this scenario (green in the overlay).
            bars = []
            for pid in (scenario.get("perspectives") or []):
                pdef = persp_defs.get(pid)
                if not pdef:
                    continue
                for b in pdef.get("bars", []):
                    pts = [_aoa_to_px(x, y) for x, y in b.get("points", [])]
                    if len(pts) >= 2:
                        bars.append({"points": pts, "label": f"{b.get('height')}cm"})

            # Minimum object size (sizePercentage = width%/height%, no position) shown as
            # a bottom-left reference box. sizePerspective (real-world cm) needs calibration
            # to scale, so it isn't drawn here.
            size_boxes = []
            for f in scenario.get("filters", []):
                if f.get("type") == "sizePercentage":
                    w, h = f.get("width", 0), f.get("height", 0)
                    if w and h:
                        wpx, hpx = int(w / 100.0 * img_w), int(h / 100.0 * img_h)
                        x0, y1 = int(0.02 * img_w), int(0.97 * img_h)
                        size_boxes.append({"box": (x0, y1 - hpx, x0 + wpx, y1), "label": "min size"})

            # Line-crossing direction (fence alarmDirection: leftToRight / rightToLeft).
            direction = triggers[0].get("alarmDirection") if triggers and triggers[0].get("type") == "fence" else None

            parsed_rules.append({"is_placeholder": False, "name": rule_name, "type": rule_type,
                                 "duration": duration_val, "target": target_detection,
                                 "vertices": vertices, "bars": bars, "size_boxes": size_boxes,
                                 "direction": direction})
        return parsed_rules


# --- UNIVERSAL RENDERING ENGINE (JPEG COMPRESSION) ---

def _draw_crossing_direction(draw, p0, p1, direction):
    """Draw a perpendicular arrow at a line's midpoint showing which way an object must
    cross to trigger. Matches the analytics writer's convention: relative to p0->p1,
    'leftToRight' points to the (-dpy, dpx) side, 'rightToLeft' the opposite; 'any'
    draws a double-headed arrow. Computed in pixel space (perpendicular to the line)."""
    import math
    x0, y0 = p0
    x1, y1 = p1
    dpx, dpy = x1 - x0, y1 - y0
    length = math.hypot(dpx, dpy) or 1.0
    if str(direction) == "rightToLeft":
        nx, ny = dpy / length, -dpx / length
    else:  # leftToRight and any
        nx, ny = -dpy / length, dpx / length
    mx, my = (x0 + x1) / 2.0, (y0 + y1) / 2.0
    arm = 42
    color = (0, 229, 218, 255)
    bidir = (str(direction) == "any")
    start = (mx - nx * arm, my - ny * arm) if bidir else (mx, my)
    tip = (mx + nx * arm, my + ny * arm)
    draw.line([start, tip], fill=color, width=3)

    def _head(at, toward):
        ang = math.atan2(at[1] - toward[1], at[0] - toward[0])
        for da in (math.radians(28), math.radians(-28)):
            draw.line([at, (at[0] + 14 * math.cos(ang + da), at[1] + 14 * math.sin(ang + da))],
                      fill=color, width=3)

    _head(tip, start)
    if bidir:
        _head(start, tip)


def render_overlay_image(camera_image, vertices, index, img_w, img_h, rule_type="",
                         size_boxes=None, bars=None, direction=None):
    snapshot_missing = (camera_image is None)
    if camera_image is not None:
        base_img = camera_image.copy().convert("RGBA")
    else:
        base_img = PILImage.new("RGBA", (img_w, img_h), (40, 40, 40, 255))

    brand_colors = [(0, 161, 154), (0, 114, 110), (229, 245, 245)]
    rgb_color = brand_colors[index % 3]

    overlay = PILImage.new("RGBA", base_img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    def _get_polygons(verts):
        if not verts:
            return []
        if isinstance(verts[0], (list, tuple)) and verts[0] and isinstance(verts[0][0], (list, tuple)):
            return verts
        return [verts]

    def _segments_intersect(a1, a2, b1, b2):
        def _orientation(p, q, r):
            return (q[1] - p[1]) * (r[0] - q[0]) - (q[0] - p[0]) * (r[1] - q[1])

        def _on_segment(p, q, r):
            return (min(p[0], r[0]) <= q[0] <= max(p[0], r[0]) and
                    min(p[1], r[1]) <= q[1] <= max(p[1], r[1]))

        o1 = _orientation(a1, a2, b1)
        o2 = _orientation(a1, a2, b2)
        o3 = _orientation(b1, b2, a1)
        o4 = _orientation(b1, b2, a2)

        if o1 != o2 and o3 != o4:
            return True
        if o1 == 0 and _on_segment(a1, b1, a2):
            return True
        if o2 == 0 and _on_segment(a1, b2, a2):
            return True
        if o3 == 0 and _on_segment(b1, a1, b2):
            return True
        if o4 == 0 and _on_segment(b1, a2, b2):
            return True
        return False

    def _polygon_is_self_intersecting(polygon):
        if len(polygon) < 4:
            return False
        n = len(polygon)
        for i in range(n):
            a1 = polygon[i]
            a2 = polygon[(i + 1) % n]
            for j in range(i + 2, n):
                if j == i or j == (i + 1) % n or (i == 0 and j == n - 1):
                    continue
                b1 = polygon[j]
                b2 = polygon[(j + 1) % n]
                if _segments_intersect(a1, a2, b1, b2):
                    return True
        return False

    def _dump_reorder_debug(original, reordered, rule_type, index, img_w, img_h):
        if os.getenv("DEBUG_OVERLAY_POLYGON", "0") != "1":
            return
        try:
            root = Path(__file__).resolve().parent
            dbg_dir = root / "debug_overlay_tests"
            dbg_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            fname = dbg_dir / f"reorder_debug_{ts}_{index}.json"
            import json
            with open(fname, "w", encoding="utf-8") as f:
                json.dump({
                    "rule_type": rule_type,
                    "index": index,
                    "img_w": img_w,
                    "img_h": img_h,
                    "original": original,
                    "reordered": reordered,
                }, f, indent=2)
        except Exception:
            pass

    for polygon in _get_polygons(vertices):
        is_line_rule = any(keyword in rule_type.lower() for keyword in ["line", "cross", "fence"])
        polygon_to_draw = polygon
        attempted_reorder = None
        if not is_line_rule and len(polygon) > 2 and _polygon_is_self_intersecting(polygon):
            try:
                import math
                cx = sum([p[0] for p in polygon]) / len(polygon)
                cy = sum([p[1] for p in polygon]) / len(polygon)
                reordered = sorted(polygon, key=lambda p: math.atan2(p[1] - cy, p[0] - cx))
                attempted_reorder = reordered
                if not _polygon_is_self_intersecting(reordered):
                    polygon_to_draw = reordered
            except Exception:
                polygon_to_draw = polygon

        if attempted_reorder is not None:
            _dump_reorder_debug(polygon, attempted_reorder, rule_type, index, img_w, img_h)

        if is_line_rule or len(polygon_to_draw) <= 2:
            # Teal outline for lines
            draw.line(polygon_to_draw, fill=(0, 161, 154, 255), width=4)
            if direction and len(polygon_to_draw) >= 2:
                _draw_crossing_direction(draw, polygon_to_draw[0], polygon_to_draw[1], direction)
        else:
            # Light red fill for polygons
            light_red_fill = (255, 102, 102, 76)
            draw.polygon(polygon_to_draw, fill=light_red_fill)
            draw.polygon(polygon_to_draw, outline=rgb_color + (255,), width=3)
        for (x, y) in polygon_to_draw:
            draw.ellipse((x - 6, y - 6, x + 6, y + 6), fill=(255, 255, 0, 255), outline=(0, 0, 0, 255), width=1)

    # Min/max object-size boxes (purple) and perspective calibration bars (green),
    # matching what the analytics writer shows. Same coordinate spaces as vertices.
    try:
        from PIL import ImageFont
        label_font = ImageFont.load_default()
    except Exception:
        label_font = None
    for sb in (size_boxes or []):
        x0, y0, x1, y1 = sb.get("box", (0, 0, 0, 0))
        draw.rectangle([x0, y0, x1, y1], outline=(177, 151, 252, 255), width=3)
        draw.text((x0 + 2, min(y0, y1) - 12), sb.get("label", ""), fill=(177, 151, 252, 255), font=label_font)
    for bar in (bars or []):
        pts = bar.get("points", [])[:2]
        if len(pts) < 2:
            continue
        draw.line([tuple(pts[0]), tuple(pts[1])], fill=(81, 207, 102, 255), width=4)
        for (bx, by) in pts:
            draw.line([(bx - 6, by), (bx + 6, by)], fill=(81, 207, 102, 255), width=3)
        draw.text((pts[0][0] + 6, pts[0][1]), bar.get("label", ""), fill=(81, 207, 102, 255), font=label_font)

    if snapshot_missing:
        try:
            from PIL import ImageFont
            font = ImageFont.load_default()
            text = "SNAPSHOT UNAVAILABLE"
            try:
                text_width, text_height = draw.textsize(text, font=font)
            except Exception:
                text_width, text_height = font.getsize(text)
            padding = 12
            text_box = [
                (base_img.width - text_width) // 2 - padding,
                (base_img.height - text_height) // 2 - padding,
                (base_img.width + text_width) // 2 + padding,
                (base_img.height + text_height) // 2 + padding,
            ]
            draw.rectangle(text_box, fill=(0, 0, 0, 200), outline=(255, 255, 255, 255), width=2)
            draw.text(((base_img.width - text_width) // 2, (base_img.height - text_height) // 2), text, fill=(255, 255, 255, 255), font=font)
        except Exception:
            draw.text((10, 10), "SNAPSHOT UNAVAILABLE", fill=(255, 255, 255, 255))

    final_img = PILImage.alpha_composite(base_img, overlay).convert("RGB")

    # Optional debug: save full-size overlay images and vertex lists when requested
    try:
        dbg_mode = os.getenv("DEBUG_OVERLAY", "0")
        if dbg_mode in ("1", "2"):
            root = Path(__file__).resolve().parent
            dbg_dir = root / "debug_overlay_tests"
            dbg_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            fname = dbg_dir / f"overlay_{ts}_{index}.png"
            final_img.save(fname, format="PNG")
            # also write vertices JSON for easier inspection
            vfname = dbg_dir / f"overlay_{ts}_{index}_vertices.json"
            try:
                import json
                with open(vfname, "w", encoding="utf-8") as vf:
                    json.dump({"vertices": vertices, "rule_type": rule_type, "img_w": img_w, "img_h": img_h}, vf, indent=2)
            except Exception:
                pass

            if dbg_mode == "2":
                # Outline-only overlay for clearer edge inspection
                try:
                    overlay_outline = PILImage.new("RGBA", base_img.size, (0, 0, 0, 0))
                    draw_outline = ImageDraw.Draw(overlay_outline)
                    if vertices:
                        draw_outline.polygon(vertices, outline=rgb_color + (255,), width=6)
                        for (x, y) in vertices:
                            draw_outline.ellipse((x - 6, y - 6, x + 6, y + 6), fill=(255, 255, 0, 255), outline=(0, 0, 0, 255), width=1)
                    img_outline = PILImage.alpha_composite(base_img, overlay_outline).convert("RGB")
                    fname_o = dbg_dir / f"overlay_{ts}_{index}_outline.png"
                    img_outline.save(fname_o, format="PNG")
                except Exception:
                    pass

                # Reordered vertices (centroid-angle sort) to see if ordering fixes shape
                try:
                    if vertices:
                        import math, json as _json
                        cx = sum([p[0] for p in vertices]) / len(vertices)
                        cy = sum([p[1] for p in vertices]) / len(vertices)
                        verts_sorted = sorted(vertices, key=lambda p: math.atan2(p[1] - cy, p[0] - cx))
                        # save reordered overlay
                        overlay_re = PILImage.new("RGBA", base_img.size, (0, 0, 0, 0))
                        draw_re = ImageDraw.Draw(overlay_re)
                        draw_re.polygon(verts_sorted, fill=rgb_color + (76,))
                        draw_re.polygon(verts_sorted, outline=rgb_color + (255,), width=3)
                        for (x, y) in verts_sorted:
                            draw_re.ellipse((x - 6, y - 6, x + 6, y + 6), fill=(255, 255, 0, 255), outline=(0, 0, 0, 255), width=1)
                        img_re = PILImage.alpha_composite(base_img, overlay_re).convert("RGB")
                        fname_r = dbg_dir / f"overlay_{ts}_{index}_reordered.png"
                        img_re.save(fname_r, format="PNG")
                        # write reordered vertices json
                        vfname_r = dbg_dir / f"overlay_{ts}_{index}_reordered_vertices.json"
                        with open(vfname_r, "w", encoding="utf-8") as vf:
                            _json.dump({"vertices": verts_sorted, "rule_type": rule_type, "img_w": img_w, "img_h": img_h}, vf, indent=2)
                except Exception:
                    pass
    except Exception:
        pass

    ratio = THUMB_H / final_img.height
    img_resized = final_img.resize((int(final_img.width * ratio), THUMB_H), PILImage.Resampling.LANCZOS)

    img_buf = BytesIO()
    img_resized.save(img_buf, format="JPEG", quality=85)
    img_buf.seek(0)

    return img_buf


# Shared layout for the per-location analytics sheets.
MAIN_HEADERS = ["Client Name", "Location", "Live Unit Serial", "Camera Position",
                "Rule Name", "Rule Type", "Target Detection", "Duration (s)",
                "Rule Visual Overlay Thumbnail"]
MAIN_WIDTHS = [22, 24, 22, 16, 20, 20, 20, 14, 95]


def _safe_sheet_title(name, used_titles):
    """Excel-worksheet-title-safe version of a location name: strip the characters
    Excel forbids (: \\ / ? * [ ]), collapse whitespace, cap at 31 chars, never
    blank, and make it unique within the workbook (append ' (2)', ' (3)', ... on a
    case-insensitive collision, still <=31). used_titles is mutated with the result."""
    cleaned = name or "Unspecified Location"
    for ch in ":\\/?*[]":
        cleaned = cleaned.replace(ch, " ")
    cleaned = " ".join(cleaned.split()).strip("'") or "Unspecified Location"
    base = cleaned[:31]
    title, n = base, 2
    while title.casefold() in used_titles:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)].rstrip() + suffix
        n += 1
    used_titles.add(title.casefold())
    return title


def _style_analytics_sheet(ws, location_label, timeline_text):
    """Apply the master-report header/styling to a per-location analytics sheet."""
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1A1D27")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                  top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Intelligent Analytics Master Report - {location_label}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color="0F1117")
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:I2")
    ws["A2"] = timeline_text
    ws["A2"].font      = Font(name="Arial", size=10, color="888888")
    ws["A2"].fill      = PatternFill("solid", start_color="0F1117")
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 20

    for col, h in enumerate(MAIN_HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        cell.border = thin
    ws.row_dimensions[3].height = 22

    for i, w in enumerate(MAIN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return ws


def add_location_sheet(wb, location_label, timeline_text, used_titles):
    """Create and style a new analytics worksheet for one location, returning it.
    The tab title is sanitized/uniquified for Excel via _safe_sheet_title."""
    ws = wb.create_sheet(title=_safe_sheet_title(location_label, used_titles))
    return _style_analytics_sheet(ws, location_label, timeline_text)


def create_master_workbook():
    """Build the workbook with just the global 'Missed Cameras' sheet. The analytics
    data is split into one tab per unique location, and those sheets are created on
    demand during the batch run via add_location_sheet() -- so this no longer creates
    a single 'Camera Analytics' sheet up front."""
    wb = openpyxl.Workbook()
    default_ws = wb.active  # openpyxl always seeds one sheet; we don't want it

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    failed_fill  = PatternFill("solid", start_color="A61C1C")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_missed = wb.create_sheet(title="Missed Cameras")
    ws_missed.views.sheetView[0].showGridLines = True

    ws_missed.merge_cells("A1:G1")
    ws_missed["A1"] = "Analytics Fetch Exception Audit Log"
    ws_missed["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_missed["A1"].fill      = PatternFill("solid", start_color="330000")
    ws_missed["A1"].alignment = center_align
    ws_missed.row_dimensions[1].height = 36

    missed_headers = ["Timestamp", "Client Name", "Location", "Live Unit Serial", "Camera Port", "Target Endpoint", "Failure Reason"]
    for col, h in enumerate(missed_headers, 1):
        cell = ws_missed.cell(row=2, column=col, value=h)
        cell.font = header_font; cell.fill = failed_fill; cell.alignment = center_align
    ws_missed.row_dimensions[2].height = 22
    missed_widths = [22, 22, 24, 22, 15, 65, 45]
    for i, w in enumerate(missed_widths, 1):
        ws_missed.column_dimensions[get_column_letter(i)].width = w

    ws_missed.freeze_panes = "A3"

    wb.remove(default_ws)  # safe now that ws_missed exists as the workbook's sheet
    return wb, ws_missed


def dedupe_main_rows(main_rows):
    """Remove exact duplicate analytics rows while preserving order."""
    seen = set()
    deduped = []
    for row in main_rows:
        key = tuple(row["data"] + [row.get("err", "")])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


# --- MULTI-THREADED WORKER FUNCTION ---

def process_camera_row(args):
    row_idx, row_data, sess, credentials = args
    client_name = row_data.get("CLIENT_NM", "").strip()
    location = row_data.get("LOCATION_NM", "").strip()
    serial = row_data.get("LIVE_UNIT_SERIAL_NM", "").strip()
    ip = row_data.get("IP", "").strip()

    user_axis, pass_axis = credentials.get("AXIS_USER"), credentials.get("AXIS_PASS")
    user_hik, pass_hik = credentials.get("HIK_USER"), credentials.get("HIK_PASS")

    raw_mfg = row_data.get("MANUFACTURER", "")
    mfg_class = classify_manufacturer(raw_mfg)
    api_target_str = mfg_class or "UNKNOWN"

    results = {
        "ip": ip, "idx": row_idx, "client": client_name, "loc": location,
        "target": api_target_str, "logs": [], "main": [], "missed": []
    }

    # Helper to append main rows while detecting duplicates during processing.
    _main_seen = set()
    def add_main_row(row):
        key = tuple(row.get("data", []) + [row.get("err", "")])
        if key in _main_seen:
            results["logs"].append(f"[DUPLICATE] Detected duplicate main row: {row.get('data')}")
        else:
            _main_seen.add(key)
        results["main"].append(row)

    if mfg_class in ("MIXED", None):
        results["logs"].append(
            f"[!] Manufacturer raw value: '{raw_mfg}' classified as {mfg_class or 'UNRECOGNIZED'}"
        )

    if not ip:
        return results

    if mfg_class is None:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        results["missed"].append([timestamp, client_name, location, serial, "N/A", "N/A",
                                   f"Unrecognized MANUFACTURER value: '{raw_mfg}' (must contain 'axis' or 'hik') -- "
                                   f"analytics skipped, snapshot attempted with both known API formats"])

        probe_handlers = []
        if user_axis and pass_axis:
            probe_handlers.append(AxisHandler(ip, user_axis, pass_axis))
        if user_hik and pass_hik:
            probe_handlers.append(HikvisionHandler(ip, user_hik, pass_hik))

        for cam in CAMERA_CONFIGS:
            port = cam["port"]
            pos = cam["position"]
            bg_color = cam["color"]
            results["logs"].append(f" -> Testing {pos} Interface Port: {port} (unrecognized manufacturer, snapshot-only)...")

            camera_image, img_w, img_h = None, 1280, 720
            for probe in probe_handlers:
                camera_image, snap_url, snap_err, _ = probe.fetch_snapshot(sess, port)
                if camera_image is not None:
                    img_w, img_h = camera_image.size
                    break
            if camera_image is None:
                results["logs"].append(f"    [!] Warning: Could not fetch snapshot with any known API format. Defaulting to {img_w}x{img_h} canvas.")

            img_buf = render_overlay_image(camera_image, [], 0, img_w, img_h, "")
            add_main_row({
                "data": [client_name, location, serial, pos, "Unrecognized Manufacturer", "N/A", "Analytics Skipped", "N/A"],
                "bg": bg_color, "img": img_buf, "err": ""
            })
            if camera_image is not None:
                camera_image.close()

        return results

    if mfg_class == "MIXED":
        axis_handler = AxisHandler(ip, user_axis, pass_axis) if (user_axis and pass_axis) else None
        hik_handler = HikvisionHandler(ip, user_hik, pass_hik) if (user_hik and pass_hik) else None

        axis_dead = axis_handler is None
        hik_dead = hik_handler is None
        center_axis_only = None

        for cam in CAMERA_CONFIGS:
            port = cam["port"]
            pos = cam["position"]
            bg_color = cam["color"]
            results["logs"].append(f" -> Testing {pos} Interface Port: {port} (mixed-vendor unit, probing)...")

            if pos == "CENTER" and axis_handler and hik_handler and center_axis_only is None:
                axis_probe_ok = False
                hik_probe_ok = False

                probe_img, _, _, probe_auth_rejected = axis_handler.fetch_snapshot(sess, port)
                if probe_img is not None:
                    probe_img.close()
                    axis_probe_ok = True
                elif probe_auth_rejected:
                    axis_dead = True

                probe_img, _, _, probe_auth_rejected = hik_handler.fetch_snapshot(sess, port)
                if probe_img is not None:
                    probe_img.close()
                    hik_probe_ok = True
                elif probe_auth_rejected:
                    hik_dead = True

                if axis_probe_ok and not hik_probe_ok:
                    center_axis_only = True
                    results["logs"].append("    [>] Mixed unit center port is Axis-only; skipping Axis probes on non-center ports.")
                elif hik_probe_ok and not axis_probe_ok:
                    center_axis_only = False
                    results["logs"].append("    [>] Mixed unit center port is Hikvision-only; non-center ports will still be probed normally.")

            handler, vendor_label, camera_image = None, None, None
            probe_errors = []

            try_axis = not axis_dead and not (center_axis_only is True and pos != "CENTER")
            if try_axis:
                img, snap_url, snap_err, auth_rejected = axis_handler.fetch_snapshot(sess, port)
                if img is not None:
                    handler, vendor_label, camera_image = axis_handler, "Axis", img
                else:
                    if auth_rejected:
                        axis_dead = True
                        results["logs"].append(f"    [!] Axis credentials flatly rejected on Port {port} -- won't retry Axis on this device's remaining ports.")
                    probe_errors.append(f"Axis: {snap_err}")
            elif center_axis_only is True:
                results["logs"].append(f"    [>] Skipping Axis probe on Port {port} because center port is Axis-only.")

            if handler is None and not hik_dead:
                img, snap_url, snap_err, auth_rejected = hik_handler.fetch_snapshot(sess, port)
                if img is not None:
                    handler, vendor_label, camera_image = hik_handler, "Hikvision", img
                else:
                    if auth_rejected:
                        hik_dead = True
                        results["logs"].append(f"    [!] Hikvision credentials flatly rejected on Port {port} -- won't retry Hikvision on this device's remaining ports.")
                    probe_errors.append(f"Hikvision: {snap_err}")

            if handler is None:
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                reason = " | ".join(probe_errors) if probe_errors else "no credentials provided for either vendor"
                results["missed"].append([timestamp, client_name, location, serial, f"{pos} ({port}) [MIXED]", "N/A",
                                           f"Neither vendor produced a snapshot on this port -- {reason}"])
                img_buf = render_overlay_image(None, [], 0, 1280, 720, "")
                add_main_row({
                        "data": [client_name, location, serial, pos, "Unresolved Vendor", "N/A", "Neither Vendor Responded", "N/A"],
                        "bg": bg_color, "img": img_buf, "err": ""
                    })
                continue

            img_w, img_h = camera_image.size
            payload_data, req_url, err_msg, analytics_auth_rejected = handler.fetch_analytics(sess, port)
            if analytics_auth_rejected:
                if vendor_label == "Axis":
                    axis_dead = True
                else:
                    hik_dead = True

            if vendor_label == "Hikvision" and payload_data is not None:
                channel_id = _extract_hik_channel_id(req_url)
                if channel_id and handler.reposition_for_scene(sess, port, channel_id, "1"):
                    results["logs"].append(f"    [+] Repositioned PTZ to analytics scene, re-fetching snapshot (Port {port}, Channel {channel_id}).")
                    fresh_img, _, _, _ = handler.fetch_snapshot(sess, port)
                    if fresh_img is not None:
                        camera_image.close()
                        camera_image = fresh_img
                        img_w, img_h = camera_image.size

            pos_label = f"{pos} [{vendor_label}]"

            if payload_data is None:
                err_msg = err_msg or "Unauthorized Connection (All auth variations failed)"
                is_special_case = err_msg.startswith("SPECIAL CASE:")
                is_no_rules = err_msg.startswith("NO_RULES:")
                if is_no_rules:
                    # Camera answered fine but has no analytic zones configured. Not a
                    # miss -- render the snapshot with a "no scenarios" placeholder.
                    placeholder_name = "No Scenarios Configured"
                    results["logs"].append(f"    [>] Port {port} responded with no analytics configured -- logging snapshot only (not a miss).")
                else:
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    tag = "[SPECIAL CASE]" if is_special_case else ""
                    results["missed"].append([timestamp, client_name, location, serial, f"{pos_label} ({port}) {tag}".strip(), req_url, err_msg])
                    placeholder_name = "Non-Standard Analytics App (See Missed Tab)" if is_special_case else "Analytics Fetch Failed (Check Missed Tab)"
                rules = [{"is_placeholder": True, "name": placeholder_name, "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]
            else:
                rules = handler.parse_analytics(payload_data, img_w, img_h)

            for index, rule in enumerate(rules):
                try:
                    img_buf = render_overlay_image(camera_image, rule["vertices"], index, img_w, img_h, rule["type"],
                                                       size_boxes=rule.get("size_boxes"), bars=rule.get("bars"), direction=rule.get("direction"))
                    display_name = rule["name"]
                    if camera_image is None:
                        display_name = f"Snapshot Failed: {display_name}"
                    add_main_row({
                        "data": [client_name, location, serial, pos_label, display_name, rule["type"], rule["target"], rule["duration"]],
                        "bg": bg_color, "img": img_buf, "err": ""
                    })
                    if rule.get("is_placeholder"):
                        results["logs"].append(f"    [+] Logged snapshot row for Port {port} (No rules active or fetch failed) -- resolved as {vendor_label}")
                    else:
                        results["logs"].append(f"    [+] Logged metrics for Port {port} Scenario: {rule['name']} ({rule['duration']}s) -- resolved as {vendor_label}")
                except Exception as e:
                    add_main_row({
                        "data": [client_name, location, serial, pos_label, rule["name"], rule["type"], rule["target"], rule["duration"]],
                        "bg": bg_color, "img": None, "err": f"(Image failed: {e})"
                    })

            if camera_image is not None:
                camera_image.close()

        return results

    is_axis = (mfg_class == "AXIS")
    handler = AxisHandler(ip, user_axis, pass_axis) if is_axis else HikvisionHandler(ip, user_hik, pass_hik)

    for cam in CAMERA_CONFIGS:
        port = cam["port"]
        pos = cam["position"]
        bg_color = cam["color"]
        results["logs"].append(f" -> Testing {pos} Interface Port: {port}...")

        if is_axis:
            # --- AXIS SPECIFIC: SNAPSHOT FIRST ---
            # Capture background snapshot immediately so auth-restricted analytics cannot block it
            camera_image, snap_url, snap_err, snap_auth_rejected = handler.fetch_snapshot(sess, port)
            if camera_image is None:
                img_w, img_h = handler.fallback_dim
                results["logs"].append(f"    [!] Warning: Failed to fetch stream snapshot ({snap_err}). Defaulting to {img_w}x{img_h} canvas.")
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                results["missed"].append([timestamp, client_name, location, serial, f"{pos} ({port}) [SNAPSHOT]", snap_url, snap_err])
            else:
                img_w, img_h = camera_image.size

            payload_data, req_url, err_msg, analytics_auth_rejected = handler.fetch_analytics(sess, port)
            if payload_data is not None and camera_image is None:
                results["logs"].append("    [!] Analytics retrieved but snapshot unavailable; rendering overlay on placeholder canvas.")
        else:
            # --- HIKVISION SPECIFIC: ANALYTICS FIRST ---
            payload_data, req_url, err_msg, analytics_auth_rejected = handler.fetch_analytics(sess, port)
            
            if payload_data is not None:
                channel_id = _extract_hik_channel_id(req_url)
                if channel_id:
                    if handler.reposition_for_scene(sess, port, channel_id, "1"):
                        results["logs"].append(f"    [+] Repositioned PTZ to analytics scene before snapshot (Port {port}, Channel {channel_id}).")

            # Fetch snapshot: fetch_snapshot will try Channel 201 first and fallback to 101 automatically
            camera_image, snap_url, snap_err, snap_auth_rejected = handler.fetch_snapshot(sess, port)
            if camera_image is None:
                img_w, img_h = handler.fallback_dim
                results["logs"].append(f"    [!] Warning: Failed to fetch stream snapshot ({snap_err}). Defaulting to {img_w}x{img_h} canvas.")
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                results["missed"].append([timestamp, client_name, location, serial, f"{pos} ({port}) [SNAPSHOT]", snap_url, snap_err])
            else:
                img_w, img_h = camera_image.size

            if payload_data is not None and camera_image is None:
                results["logs"].append("    [!] Analytics retrieved but snapshot unavailable; rendering overlay on placeholder canvas.")

        if payload_data is None:
            err_msg = err_msg or "Unauthorized Connection (All auth variations failed)"
            is_special_case = err_msg.startswith("SPECIAL CASE:")
            is_no_rules = err_msg.startswith("NO_RULES:")
            if is_no_rules:
                # Camera answered fine but has no analytic zones configured. Not a
                # miss -- render the snapshot with a "no scenarios" placeholder.
                placeholder_name = "No Scenarios Configured"
                results["logs"].append(f"    [>] Port {port} responded with no analytics configured -- logging snapshot only (not a miss).")
            else:
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                tag = "[SPECIAL CASE]" if is_special_case else ""
                results["missed"].append([timestamp, client_name, location, serial, f"{pos} ({port}) {tag}".strip(), req_url, err_msg])
                placeholder_name = "Non-Standard Analytics App (See Missed Tab)" if is_special_case else "Analytics Fetch Failed (Check Missed Tab)"
            rules = [{"is_placeholder": True, "name": placeholder_name, "type": "N/A", "target": "No Analytics Configured", "duration": "N/A", "vertices": []}]
        else:
            rules = handler.parse_analytics(payload_data, img_w, img_h)

        for index, rule in enumerate(rules):
            try:
                img_buf = render_overlay_image(camera_image, rule["vertices"], index, img_w, img_h, rule["type"],
                                                       size_boxes=rule.get("size_boxes"), bars=rule.get("bars"), direction=rule.get("direction"))
                add_main_row({
                    "data": [client_name, location, serial, pos, rule["name"], rule["type"], rule["target"], rule["duration"]],
                    "bg": bg_color, "img": img_buf, "err": ""
                })
                if rule.get("is_placeholder"):
                    results["logs"].append(f"    [+] Logged snapshot row for Port {port} (No rules active or fetch failed)")
                else:
                    results["logs"].append(f"    [+] Logged metrics for Port {port} Scenario: {rule['name']} ({rule['duration']}s)")
            except Exception as e:
                add_main_row({
                    "data": [client_name, location, serial, pos, rule["name"], rule["type"], rule["target"], rule["duration"]],
                    "bg": bg_color, "img": None, "err": f"(Image failed: {e})"
                })

        if camera_image is not None:
            camera_image.close()

        # --- DECOUPLED ACCOUNT LOCKOUT safeguard EVALUATION ---
        # For Axis, only trigger loop break on direct snapshot 401 refusal.
        # For Hikvision, preserve both checks intact to prevent lockout windows.
        should_break = snap_auth_rejected if is_axis else (snap_auth_rejected or analytics_auth_rejected)
        if should_break:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            results["logs"].append(f"    [!] Authentication flatly rejected on Port {port} -- skipping remaining ports on this device to avoid triggering an account lockout.")
            results["missed"].append([timestamp, client_name, location, serial, "REMAINING PORTS SKIPPED", "N/A",
                                       f"Authentication rejected on port {port} -- remaining ports skipped to avoid repeated failed-login attempts against the same device."])
            break

    return results


def run_batch(camera_rows, credentials, output_dir, base_filename, log_cb=None, progress_cb=None):
    """Runs the full batch: spins up the thread pool, processes every row, writes
    the Excel report, and returns the saved Path.

    credentials: dict with any of AXIS_USER/AXIS_PASS/HIK_USER/HIK_PASS.
    log_cb(str): called for every line of progress output (defaults to a no-op).
    progress_cb(done, total): called after each row finishes (defaults to a no-op).
    """
    log = log_cb or (lambda _msg: None)
    report_progress = progress_cb or (lambda _done, _total: None)

    script_start_time = time.time()
    date_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    master_report_file = f"{base_filename}_Master_{date_suffix}.xlsx"

    current_missed_row = 3

    wb, ws_missed = create_master_workbook()

    # One analytics tab per unique location, created on demand. Each entry tracks
    # its worksheet and the next free row on that sheet (data starts at row 4).
    timeline_text = f"Batch Processing Timeline Context: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    loc_sheets = {}
    used_titles = set()

    def get_location_sheet(loc):
        key = (loc or "").strip() or "Unspecified Location"
        if key not in loc_sheets:
            ws = add_location_sheet(wb, key, timeline_text, used_titles)
            loc_sheets[key] = {"ws": ws, "row": 4}
        return loc_sheets[key]

    normal_font = Font(name="Arial", size=10)
    bold_font   = Font(name="Arial", bold=True, size=10)
    white_normal_font = Font(name="Arial", size=10, color="FFFFFF")
    white_bold_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    left_align  = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"), top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

    session = requests.Session()
    retries = Retry(total=2, backoff_factor=0.5, status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries, pool_connections=MAX_WORKER_THREADS, pool_maxsize=MAX_WORKER_THREADS)
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    total_cameras = len(camera_rows)
    log("[+] Spinning up multi-threaded batch engine...")

    active_worker_count = max(1, min(MAX_WORKER_THREADS, total_cameras))

    done_count = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=active_worker_count) as executor:
        worker_args = [(i, row, session, credentials) for i, row in enumerate(camera_rows, 1)]

        for res in executor.map(process_camera_row, worker_args):
            if not res["ip"]:
                done_count += 1
                report_progress(done_count, total_cameras)
                continue

            log(f"\n#################################################################")
            log(f" INGESTING HOST NODE [{res['idx']}/{total_cameras}]: {res['client']} - {res['loc']} ({res['ip']}) [{res['target']}]")
            log(f"#################################################################")
            for line in res["logs"]:
                log(line)

            for missed_row in res["missed"]:
                ws_missed.append(missed_row)
                for col in range(1, 8):
                    ws_missed.cell(row=current_missed_row, column=col).border = thin_border
                    ws_missed.cell(row=current_missed_row, column=col).font = normal_font
                current_missed_row += 1

            main_rows = dedupe_main_rows(res["main"])
            if len(main_rows) < len(res["main"]):
                log(f"[!] Removed {len(res['main']) - len(main_rows)} duplicate analytics row(s) for IP {res['ip']}")
            if main_rows:
                entry = get_location_sheet(res["loc"])
                ws_loc = entry["ws"]
            for main_row in main_rows:
                current_main_row = entry["row"]
                bg = main_row["bg"]
                row_fill = PatternFill("solid", start_color=bg)
                active_normal_font = white_normal_font if bg == "00726E" else normal_font
                active_bold_font = white_bold_font if bg == "00726E" else bold_font

                for col_idx, val in enumerate(main_row["data"], 1):
                    c = ws_loc.cell(row=current_main_row, column=col_idx, value=val)
                    c.font = active_bold_font if col_idx == 4 else active_normal_font
                    c.fill = row_fill
                    c.alignment = center_align if col_idx in [3, 4, 7, 8] else left_align
                    c.border = thin_border

                thumb_cell = ws_loc.cell(row=current_main_row, column=9, value=main_row["err"])
                thumb_cell.fill = row_fill; thumb_cell.border = thin_border

                if main_row["img"]:
                    xl_img = OpenpyxlImage(main_row["img"])
                    xl_img.anchor = TwoCellAnchor(editAs="oneCell", _from=AnchorMarker(col=8, colOff=0, row=current_main_row - 1, rowOff=0), to=AnchorMarker(col=9, colOff=0, row=current_main_row, rowOff=0))
                    ws_loc.add_image(xl_img)

                ws_loc.row_dimensions[current_main_row].height = ROW_H
                entry["row"] = current_main_row + 1

            done_count += 1
            report_progress(done_count, total_cameras)

    # Keep the global Missed Cameras tab last so the report opens on a location tab.
    missed = wb["Missed Cameras"]
    wb._sheets.remove(missed)
    wb._sheets.append(missed)

    final_output_path = Path(output_dir) / master_report_file
    wb.save(final_output_path)

    elapsed_seconds = time.time() - script_start_time
    minutes = int(elapsed_seconds // 60)
    seconds = int(elapsed_seconds % 60)

    log(f"\n=================================================================")
    log(f" MASTER BATCH PROCESSING COMPLETED SUCCESSFULLY")
    log(f"=================================================================")
    log(f" -> Total Run Duration: {minutes}m {seconds}s ({round(elapsed_seconds, 2)} total seconds)")
    log(f" -> Threads Utilized: {active_worker_count}")
    log(f" -> Master Report Saved: {final_output_path.name}")
    log(f"=================================================================\n")

    return final_output_path