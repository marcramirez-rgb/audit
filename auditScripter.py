import os
import time
import platform
import subprocess
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Start execution timer
script_start_time = time.time()

downloads_path = Path.home() / "Downloads"
output_dir = downloads_path / "Uptime_Analysis_Reports"
output_dir.mkdir(parents=True, exist_ok=True)

# Style Tokens
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1A1D27")
zebra_fill = PatternFill("solid", start_color="F7F9FA")
white_fill = PatternFill("solid", start_color="FFFFFF")
positive_fill = PatternFill("solid", start_color="E6F4EA") # Soft green
negative_fill = PatternFill("solid", start_color="FCE8E6") # Soft red

left_align = Alignment(horizontal="left", vertical="center")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD")
)

def normalize_string(val):
    """Normalizes headers to catch spelling/case inconsistencies."""
    if val is None:
        return ""
    return str(val).strip().lower().replace("_", "").replace(" ", "")

def extract_sheet_data(file_path):
    """Scrapes Excel sheet dynamically identifying all target metadata columns."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    serial_col_idx = None
    uptime_col_idx = None
    client_col_idx = None
    model_col_idx = None
    
    # Scan Header Row (Scanning first 5 rows to locate structural positions)
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            cell_val = normalize_string(ws.cell(row=r, column=c).value)
            if "serial" in cell_val:
                serial_col_idx = c
            elif "uptime" in cell_val or "percentage" in cell_val:
                uptime_col_idx = c
            elif "parent" in cell_val:  # Target Parent Client Name explicitly
                client_col_idx = c
            elif "model" in cell_val or "equipment" in cell_val or "generator" in cell_val:
                model_col_idx = c
                
        if serial_col_idx and uptime_col_idx:
            start_row = r + 1
            break
    else:
        # Fallback defaults if manual pattern scanner misses structural layout
        serial_col_idx, uptime_col_idx, client_col_idx, model_col_idx, start_row = 5, 7, 1, 6, 2

    data_matrix = {}
    last_known_client = "Unknown Client"  # Carry-over variable for sparse/blank cells
    
    for row in range(start_row, ws.max_row + 1):
        serial = str(ws.cell(row=row, column=serial_col_idx).value or "").strip()
        uptime_raw = str(ws.cell(row=row, column=uptime_col_idx).value or "0").strip()
        
        # Read parent client cell; if it's empty, use the last known valid parent client name
        client_raw = ws.cell(row=row, column=client_col_idx).value if client_col_idx else None
        if client_raw and str(client_raw).strip() != "" and str(client_raw).strip() != "None":
            last_known_client = str(client_raw).strip()
            
        model_val = str(ws.cell(row=row, column=model_col_idx).value or "No EFOY").strip() if model_col_idx else "No EFOY"
        
        if not serial or serial == "None" or "master" in serial.lower():
            continue
            
        # Strip percentage markers to sanitize mathematics payload
        uptime_clean = uptime_raw.replace("%", "").strip()
        try:
            uptime_float = float(uptime_clean)
        except ValueError:
            uptime_float = 0.0
            
        data_matrix[serial] = {
            "uptime": uptime_float,
            "client": last_known_client,
            "model": model_val
        }
        
    wb.close()
    return data_matrix

# --- 1. FILE SELECTION INTERACTION RUNTIME ---
print("=========================================")
root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)

print("[*] Please select your Baseline File (e.g., Q1 or Older Data)...")
file1_path = filedialog.askopenfilename(initialdir=downloads_path, title="Select Baseline File (File 1)", filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))

print("[*] Please select your Comparison File (e.g., Q2 or Newer Data)...")
file2_path = filedialog.askopenfilename(initialdir=downloads_path, title="Select Comparison File (File 2)", filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))

if not file1_path or not file2_path:
    print("[-] User cancel interaction triggered. Missing files. Exiting.")
    exit()

# --- 2. MULTI-METADATA EXTRACTION ENGINE ---
print("\n[+] Scraping comprehensive historical files...")
q1_data = extract_sheet_data(file1_path)
q2_data = extract_sheet_data(file2_path)

all_serials = sorted(list(set(q1_data.keys()) | set(q2_data.keys())))

# Pre-compile data into a structured list so we can sort dynamically
compiled_rows = []
for serial in all_serials:
    node_q1 = q1_data.get(serial, {})
    node_q2 = q2_data.get(serial, {})
    
    u1 = node_q1.get("uptime", 0.0)
    u2 = node_q2.get("uptime", 0.0)
    delta = u2 - u1
    
    client = node_q2.get("client") or node_q1.get("client") or "N/A"
    model = node_q2.get("model") or node_q1.get("model") or "No EFOY"
    
    compiled_rows.append({
        "client": client,
        "model": model,
        "serial": serial,
        "u1": u1,
        "u2": u2,
        "delta": delta
    })

# SORT BY BASELINE UPTIME (u1) DESCENDING
compiled_rows.sort(key=lambda x: x["u1"], reverse=True)


# --- 3. EXCEL COMPILING GENERATION ---
out_wb = openpyxl.Workbook()

ws_data = out_wb.active
ws_data.title = "Uptime Analytics Data"
ws_data.views.sheetView[0].showGridLines = True

# Extended Header Span Banner
ws_data.merge_cells("A1:F1")
ws_data["A1"] = "Quarterly System Performance & Variance Master Matrix"
ws_data["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws_data["A1"].fill = PatternFill("solid", start_color="0F1117")
ws_data["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_data.row_dimensions[1].height = 35

# Expanded column layout configuration
headers = ["Parent Client Name", "Generator Model", "Live Unit Serial", "Baseline Uptime (%)", "Comparison Uptime (%)", "Variance Shift (Delta)"]
for idx, h in enumerate(headers, 1):
    cell = ws_data.cell(row=2, column=idx, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align

# Populate records block loop using sorted list data
current_row = 3
for item in compiled_rows:
    row_fill = zebra_fill if current_row % 2 == 0 else white_fill
    delta_fill = positive_fill if item["delta"] > 0 else (negative_fill if item["delta"] < 0 else row_fill)
    
    c1 = ws_data.cell(row=current_row, column=1, value=item["client"])
    c2 = ws_data.cell(row=current_row, column=2, value=item["model"])
    c3 = ws_data.cell(row=current_row, column=3, value=item["serial"])
    
    # Adjust values to fractional percentages for correct formatting
    c4 = ws_data.cell(row=current_row, column=4, value=item["u1"] / 100.0 if item["u1"] > 1.0 else item["u1"])
    c5 = ws_data.cell(row=current_row, column=5, value=item["u2"] / 100.0 if item["u2"] > 1.0 else item["u2"])
    c6 = ws_data.cell(row=current_row, column=6, value=item["delta"] / 100.0 if abs(item["delta"]) > 1.0 else item["delta"])
    
    for c_idx, cell in enumerate([c1, c2, c3, c4, c5, c6], 1):
        cell.border = thin_border
        cell.fill = delta_fill if c_idx == 6 else row_fill
        cell.alignment = left_align if c_idx in [1, 2, 3] else center_align
        
        if c_idx > 3:
            cell.number_format = '0.0%'
            
    ws_data.row_dimensions[current_row].height = 20
    current_row += 1

# Professional Workbook Width Adjustments
ws_data.column_dimensions['A'].width = 32 
ws_data.column_dimensions['B'].width = 22 
ws_data.column_dimensions['C'].width = 24 
ws_data.column_dimensions['D'].width = 20 
ws_data.column_dimensions['E'].width = 22 
ws_data.column_dimensions['F'].width = 22 

# --- 4. EMBEDDING DYNAMIC CHART TAB ---
ws_chart = out_wb.create_sheet(title="Uptime Delta Chart")
ws_chart.views.sheetView[0].showGridLines = True

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Quarterly Unit Uptime Shifts (Sorted by Baseline Performance)"
chart.y_axis.title = "Uptime Delta Change (%)"
chart.x_axis.title = "Device Unique Identifier (Serial)"
chart.y_axis.number_format = '0%'

# Shift reference parameters over to match Delta Column 6
data_ref = Reference(ws_data, min_col=6, min_row=2, max_row=current_row-1)
# Category titles mapped to Serial Column 3
cats_ref = Reference(ws_data, min_col=3, min_row=3, max_row=current_row-1)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.legend = None

chart.height = 15
chart.width = 26
ws_chart.add_chart(chart, "B3")

# --- 5. EXPORT ENGINE ---
timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
final_output_path = output_dir / f"Uptime_Quarterly_Matched_Analysis_{timestamp_str}.xlsx"
out_wb.save(final_output_path)

elapsed_seconds = time.time() - script_start_time
print(f"\n=================================================================")
print(f" COMPREHENSIVE ANALYSIS METRIC MATRIX RUN COMPLETE")
print(f"=================================================================")
print(f" -> Total Managed Nodes Unified: {len(all_serials)}")
print(f" -> Execution Run Time: {round(elapsed_seconds, 2)} seconds")
print(f" -> Saved Master Report: {final_output_path.name}")
print(f"=================================================================\n")

try:
    folder_to_open = final_output_path.parent
    if platform.system() == "Windows":
        os.startfile(folder_to_open)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", folder_to_open])
    else:
        subprocess.Popen(["xdg-open", folder_to_open])
except Exception:
    pass